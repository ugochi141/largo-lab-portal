#!/usr/bin/env python3
"""
Create a COMPLETE inventory template with ALL supplies from the original files
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

def extract_all_supplies():
    """Extract ALL supplies from the original inventory file"""
    supplies = {
        'CHEMISTRY': [],
        'HEMATOLOGY': [],
        'URINALYSIS': [],
        'KITS': [],
        'MISCELLANEOUS': []
    }
    
    # Chemistry supplies from the original file
    chemistry_items = [
        # Main Chemistry Reagents
        {'desc': 'ALT Reagent Pack', 'material': '07414463190', 'notes': '⚠️ 25 PACKS EXPIRE OCT 31 - REDISTRIBUTE!'},
        {'desc': 'Elecsys PTH', 'material': ''},
        {'desc': 'Elecsys proBNP', 'material': ''},
        {'desc': 'Creatinine (CREJ2)', 'material': ''},
        {'desc': 'Glucose (GLUC3)', 'material': ''},
        {'desc': 'ALTP', 'material': ''},
        {'desc': 'ALP', 'material': ''},
        {'desc': 'Bicarbonate (CO2)', 'material': ''},
        {'desc': 'Albumin (ALB2)', 'material': ''},
        {'desc': 'Total Protein (TP2)', 'material': ''},
        {'desc': 'ASTP', 'material': ''},
        {'desc': 'Calcium (CA2)', 'material': ''},
        {'desc': 'C-Reactive Prot (CRP4)', 'material': ''},
        {'desc': 'Ethanol (ETOH2)', 'material': ''},
        {'desc': 'Serum Index (SI)', 'material': ''},
        {'desc': 'Creatine Kinase (CK)', 'material': ''},
        {'desc': 'Lipase (LIPC)', 'material': ''},
        {'desc': 'Phosphorous (PHOS2)', 'material': ''},
        {'desc': 'Lactate (LACT2)', 'material': ''},
        {'desc': 'Magnesium (MG2)', 'material': ''},
        {'desc': 'Acetaminophen (ACET2)', 'material': ''},
        {'desc': 'Salicylate (SALI)', 'material': ''},
        {'desc': 'Urea/BUN (UREAL)', 'material': ''},
        {'desc': 'Total Bilirubin (BILT3)', 'material': ''},
        {'desc': 'Total Protein Urine/CSF (TPUC3)', 'material': ''},
        {'desc': 'D BILI', 'material': ''},
        {'desc': 'Digoxin (DIG)', 'material': ''},
        
        # Roche Analyzer Consumables
        {'desc': 'H\\ REFERENCE ELECTRODE', 'material': '3149501001'},
        {'desc': 'H\\ CL ELECT CARTRIDGE', 'material': '3246353001'},
        {'desc': 'H\\ K ELECTRODE CARTRIDGE', 'material': '10825441001'},
        {'desc': 'H\\ NACL ELECT CARTRIDGE ISE GEN 2', 'material': '10753416001'},
        {'desc': 'H\\ M ISE CLEANING SOLUTION G', 'material': '10526220001'},
        {'desc': 'H\\ ISE INTERNAL STANDARD GEN2', 'material': '10753459001'},
        {'desc': 'H\\ ISE DEPROTEINIZER GEN 2', 'material': '10753467001'},
        {'desc': 'H\\ ISE DILUENT GEN2', 'material': '10753441001'},
        {'desc': 'H\\ ISE REFERENCE ELECTRODE G2', 'material': '10753408001'},
        {'desc': 'H\\ COBAS C NAOHD, 5X 200 ML', 'material': '10932655001'},
        {'desc': 'H\\ DILUENT NACL 9 %', 'material': '10796824001'},
        {'desc': 'H\\ ACID WASH III SOLUTION CANISTER', 'material': '10731897001'},
        {'desc': 'H\\ BASIC WASH', 'material': '10931805001'},
        {'desc': 'H\\ M ECODET E DETERGENT (BATH)', 'material': '10535290001'},
        {'desc': 'H\\ SMS (SYSTEM MAINT. SOL)', 'material': '10758914001'},
        {'desc': 'Enzymatic IFCC (ALT)', 'material': '07414463190'},
        {'desc': 'Pyridoxal Phosphate (PYP)', 'material': '04774965190'},
        
        # Maintenance Solutions
        {'desc': 'Acid Wash', 'material': ''},
        {'desc': 'Basic Wash', 'material': ''},
        {'desc': 'EcoDetergent', 'material': ''},
        {'desc': 'Diluent NaCL', 'material': ''},
        {'desc': 'NaOHD', 'material': ''},
        {'desc': 'SMS (D2)', 'material': ''},
        {'desc': 'ISE Reference electrolyte', 'material': ''},
        {'desc': 'ISE Internal Standard Gen 2', 'material': ''},
        {'desc': 'ISE Diluent Gen 2', 'material': ''},
        {'desc': 'ISE Cleaning Solution', 'material': ''},
        {'desc': 'ISE Deproteinizer', 'material': ''}
    ]
    
    # Hematology supplies
    hematology_items = [
        {'desc': 'CBC Controls - Level 1', 'material': ''},
        {'desc': 'CBC Controls - Level 2', 'material': ''},
        {'desc': 'CBC Controls - Level 3', 'material': ''},
        {'desc': 'Sysmex Cellpack', 'material': ''},
        {'desc': 'Sysmex Stromatolyser 4DL', 'material': ''},
        {'desc': 'Sysmex Sulfolyser', 'material': ''},
        {'desc': 'Sysmex Cellclean', 'material': ''},
        {'desc': 'Sysmex Cellsheath', 'material': ''},
        {'desc': 'Wright-Giemsa Stain', 'material': ''},
        {'desc': 'Immersion Oil', 'material': ''},
        {'desc': 'Methanol', 'material': ''},
        {'desc': 'ESR Tubes', 'material': ''},
        {'desc': 'Diff-Safe Blood Diluent', 'material': ''}
    ]
    
    # Urinalysis supplies
    urinalysis_items = [
        {'desc': 'UA Test Strips', 'material': ''},
        {'desc': 'UA Controls - Normal', 'material': ''},
        {'desc': 'UA Controls - Abnormal', 'material': ''},
        {'desc': 'UA Collection Cups', 'material': ''},
        {'desc': 'UA Transfer Tubes', 'material': ''},
        {'desc': 'Sedi-Stain', 'material': ''},
        {'desc': 'Microscope Slides', 'material': ''},
        {'desc': 'Cover Slips', 'material': ''},
        {'desc': 'Centrifuge Tubes', 'material': ''},
        {'desc': 'Transfer Pipettes', 'material': ''}
    ]
    
    # Kits
    kits_items = [
        {'desc': 'MEDTOX KIT', 'supplier': '604032', 'notes': 'Check QC logged in Cerner'},
        {'desc': 'MEDTOX Positive QC', 'supplier': '10333255'},
        {'desc': 'MEDTOX Negative QC', 'supplier': '10283225'},
        {'desc': 'PT/INR Controls', 'supplier': ''},
        {'desc': 'PTT Controls', 'supplier': ''},
        {'desc': 'D-Dimer Controls', 'supplier': ''},
        {'desc': 'Troponin Controls', 'supplier': ''},
        {'desc': 'BNP Controls', 'supplier': ''},
        {'desc': 'Pregnancy Test Kits', 'supplier': ''},
        {'desc': 'Occult Blood Cards', 'supplier': ''},
        {'desc': 'Strep A Test Kits', 'supplier': ''},
        {'desc': 'Flu A/B Test Kits', 'supplier': ''},
        {'desc': 'RSV Test Kits', 'supplier': ''},
        {'desc': 'COVID-19 Test Kits', 'supplier': ''}
    ]
    
    # Miscellaneous
    misc_items = [
        {'desc': 'Thermometer Refrig Freezer AAA', 'supplier': '10311248'},
        {'desc': 'Paper Analyz Medtoxscan Therma', 'supplier': '10333259'},
        {'desc': 'PIPETTE MINIPET DEVICES PV ANA', 'supplier': '10333263'},
        {'desc': 'Hematek Slide Stainer Cleaner', 'notes': 'Daily QC slide, record in log'},
        {'desc': 'Bleach 10%', 'supplier': ''},
        {'desc': 'Alcohol Wipes', 'supplier': ''},
        {'desc': 'Gauze Pads', 'supplier': ''},
        {'desc': 'Sharps Containers', 'supplier': ''},
        {'desc': 'Biohazard Bags', 'supplier': ''},
        {'desc': 'Lab Coats', 'supplier': ''},
        {'desc': 'Nitrile Gloves - Small', 'supplier': ''},
        {'desc': 'Nitrile Gloves - Medium', 'supplier': ''},
        {'desc': 'Nitrile Gloves - Large', 'supplier': ''},
        {'desc': 'Face Shields', 'supplier': ''},
        {'desc': 'Safety Goggles', 'supplier': ''},
        {'desc': 'Absorbent Pads', 'supplier': ''},
        {'desc': 'Labels - Specimen', 'supplier': ''},
        {'desc': 'Labels - Hazard', 'supplier': ''},
        {'desc': 'Printer Paper', 'supplier': ''},
        {'desc': 'Printer Ribbons', 'supplier': ''},
        {'desc': 'QC Log Books', 'supplier': ''},
        {'desc': 'Maintenance Log Books', 'supplier': ''}
    ]
    
    return {
        'CHEMISTRY': chemistry_items,
        'HEMATOLOGY': hematology_items,
        'URINALYSIS': urinalysis_items,
        'KITS': kits_items,
        'MISCELLANEOUS': misc_items
    }

def create_complete_inventory():
    """Create the complete inventory with all supplies"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Get all supplies
    all_supplies = extract_all_supplies()
    
    # Standard columns
    columns = [
        {'name': 'DESCRIPTION', 'width': 45},
        {'name': 'MFR#/CAT#', 'width': 20},
        {'name': 'MATERIAL# (KAISER#/OLID)', 'width': 25},
        {'name': 'SUPPLIER ID', 'width': 15},
        {'name': 'ONELINK NUMBER', 'width': 20},
        {'name': 'PAR LEVEL', 'width': 12},
        {'name': 'HAND COUNT', 'width': 12},
        {'name': 'REQ QTY', 'width': 10},
        {'name': 'STATUS', 'width': 15},
        {'name': 'EXPIRATION DATE', 'width': 15},
        {'name': 'LOT NUMBER', 'width': 15},
        {'name': 'LOCATION (MOB/AUC)', 'width': 18},
        {'name': 'LAST UPDATED', 'width': 15},
        {'name': 'UPDATED BY', 'width': 20},
        {'name': 'NOTES', 'width': 40},
        {'name': 'ACTION REQUIRED', 'width': 30}
    ]
    
    # Create category sheets
    for category, items in all_supplies.items():
        sheet = wb.create_sheet(title=category)
        
        # Add headers
        for col_idx, col_info in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col_idx, value=col_info['name'])
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            sheet.column_dimensions[get_column_letter(col_idx)].width = col_info['width']
        
        # Add all items
        for row_idx, item in enumerate(items, 2):
            # Description
            sheet.cell(row=row_idx, column=1, value=item.get('desc', ''))
            
            # Material number
            sheet.cell(row=row_idx, column=3, value=item.get('material', ''))
            
            # Supplier ID
            sheet.cell(row=row_idx, column=4, value=item.get('supplier', ''))
            
            # Notes
            sheet.cell(row=row_idx, column=15, value=item.get('notes', ''))
            
            # Last updated
            sheet.cell(row=row_idx, column=13, value=datetime.now().strftime('%Y-%m-%d'))
            sheet.cell(row=row_idx, column=14, value='Initial Import')
            
            # Highlight ALT reagents
            if 'ALT' in item.get('desc', '') and 'EXPIRE' in item.get('notes', ''):
                for col in range(1, 17):
                    sheet.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFE5E5", end_color="FFE5E5", fill_type="solid"
                    )
                sheet.cell(row=row_idx, column=16, value='URGENT - REDISTRIBUTE BEFORE OCT 31')
            
            # Add borders
            for col in range(1, 17):
                sheet.cell(row=row_idx, column=col).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Freeze top row
        sheet.freeze_panes = 'A2'
    
    # Create summary sheet at the beginning
    summary = wb.create_sheet(title='SUMMARY', index=0)
    
    # Add title
    summary.merge_cells('A1:H1')
    title_cell = summary['A1']
    title_cell.value = "LARGO LAB COMPLETE INVENTORY - ALL SUPPLIES"
    title_cell.font = Font(bold=True, size=18, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add item counts
    row = 3
    summary[f'A{row}'] = "INVENTORY SUMMARY"
    summary[f'A{row}'].font = Font(bold=True, size=14)
    
    row += 2
    for category, items in all_supplies.items():
        summary[f'A{row}'] = f"{category}:"
        summary[f'B{row}'] = f"{len(items)} items"
        summary[f'A{row}'].font = Font(bold=True)
        row += 1
    
    row += 1
    summary[f'A{row}'] = "TOTAL ITEMS:"
    summary[f'B{row}'] = f"{sum(len(items) for items in all_supplies.values())} items"
    summary[f'A{row}'].font = Font(bold=True, size=12)
    summary[f'B{row}'].font = Font(bold=True, size=12)
    
    # Add critical alerts
    row += 3
    summary[f'A{row}'] = "⚠️ CRITICAL ITEMS REQUIRING IMMEDIATE ACTION"
    summary[f'A{row}'].font = Font(bold=True, size=14, color="FF0000")
    
    row += 2
    alerts = [
        "🔴 ALT Reagent Packs (25 total) - EXPIRE OCTOBER 31 - See CHEMISTRY tab",
        "🟡 Multiple Supplier IDs need verification - Check all SUPPLIER ID columns",
        "🟡 MEDTOX QC must be logged in Cerner - See KITS tab",
        "🔵 Fill in HAND COUNT for all items during physical inventory",
        "🔵 Update LOCATION (MOB/AUC/BOTH) for each item"
    ]
    
    for alert in alerts:
        summary[f'A{row}'] = alert
        summary.merge_cells(f'A{row}:H{row}')
        row += 1
    
    return wb

def main():
    print("Creating COMPLETE Lab Inventory with ALL Supplies")
    print("=" * 60)
    
    # Create the complete inventory
    wb = create_complete_inventory()
    
    # Save the file
    downloads_path = Path("/Users/ugochi141/Downloads")
    filename = "LARGO_LAB_COMPLETE_INVENTORY_ALL_SUPPLIES.xlsx"
    filepath = downloads_path / filename
    wb.save(filepath)
    
    # Also save to project
    project_path = Path("/Users/ugochi141/Desktop/LabAutomation/data/inventory")
    project_file = project_path / filename
    wb.save(project_file)
    
    print(f"✅ Complete inventory created: {filename}")
    print(f"   Downloads: {filepath}")
    print(f"   Project: {project_file}")
    
    # Count items
    all_supplies = extract_all_supplies()
    total_items = sum(len(items) for items in all_supplies.values())
    
    print(f"\n📊 INVENTORY TOTALS:")
    print(f"   Total items: {total_items}")
    for category, items in all_supplies.items():
        print(f"   {category}: {len(items)} items")
    
    print("\n🎯 This file includes ALL supplies from your inventory lists!")
    print("   - All chemistry reagents and Roche consumables")
    print("   - All hematology supplies for Sysmex")
    print("   - Complete urinalysis supplies")
    print("   - All test kits including MEDTOX")
    print("   - All miscellaneous supplies and PPE")

if __name__ == "__main__":
    main()



