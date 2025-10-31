#!/usr/bin/env python3
"""
Create a comprehensive inventory template with sample data from chat logs
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
from pathlib import Path

def create_inventory_template():
    """Create a comprehensive inventory management template"""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create a date style
    date_style = NamedStyle(name='date_style')
    date_style.number_format = 'YYYY-MM-DD'
    wb.add_named_style(date_style)
    
    # Define categories based on the lab structure
    categories = {
        'CHEMISTRY': {
            'items': [
                {'desc': 'ALT Reagent Pack', 'material': '1234567', 'supplier': '10333255', 'hand_count': '25', 'location': 'MOB/AUC', 'expiration': '2025-10-31', 'notes': '25 packs split between MOB and AUC - redistribute before expiration'},
                {'desc': 'H\\ REFERENCE ELECTRODE', 'material': '3149501001', 'supplier': '', 'hand_count': '', 'location': '', 'expiration': '', 'notes': ''},
                {'desc': 'H\\ CL ELECT CARTRIDGE', 'material': '3246353001', 'supplier': '', 'hand_count': '', 'location': '', 'expiration': '', 'notes': ''},
                {'desc': 'H\\ K ELECTRODE CARTRIDGE', 'material': '10825441001', 'supplier': '', 'hand_count': '', 'location': '', 'expiration': '', 'notes': ''},
            ]
        },
        'HEMATOLOGY': {
            'items': [
                {'desc': 'CBC Controls - Level 1', 'material': '', 'supplier': '', 'hand_count': '', 'location': 'MOB', 'expiration': '', 'notes': 'For Sysmex XN 1000'},
                {'desc': 'CBC Controls - Level 2', 'material': '', 'supplier': '', 'hand_count': '', 'location': 'MOB', 'expiration': '', 'notes': 'For Sysmex XN 1000'},
                {'desc': 'CBC Controls - Level 3', 'material': '', 'supplier': '', 'hand_count': '', 'location': 'MOB', 'expiration': '', 'notes': 'For Sysmex XN 1000'},
            ]
        },
        'URINALYSIS': {
            'items': [
                {'desc': 'UA Test Strips', 'material': '', 'supplier': '', 'hand_count': '', 'location': 'AUC', 'expiration': '', 'notes': 'Send overflow to AUC when Lorraine is out'},
                {'desc': 'UA Controls', 'material': '', 'supplier': '', 'hand_count': '', 'location': 'BOTH', 'expiration': '', 'notes': ''},
                {'desc': 'UA Collection Cups', 'material': '', 'supplier': '', 'hand_count': '', 'location': 'BOTH', 'expiration': '', 'notes': ''},
            ]
        },
        'KITS': {
            'items': [
                {'desc': 'MEDTOX KIT', 'material': '', 'supplier': '604032', 'hand_count': '', 'location': 'BOTH', 'expiration': '', 'notes': 'Check QC is logged in Cerner'},
                {'desc': 'MEDTOX Positive QC', 'material': '', 'supplier': '10333255', 'hand_count': '5', 'location': 'BOTH', 'expiration': '', 'notes': ''},
                {'desc': 'MEDTOX Negative QC', 'material': '', 'supplier': '10283225', 'hand_count': '1', 'location': 'BOTH', 'expiration': '', 'notes': ''},
            ]
        },
        'MISCELLANEOUS': {
            'items': [
                {'desc': 'Thermometer Refrig Freezer AAA', 'material': '', 'supplier': '10311248', 'hand_count': '', 'location': '', 'expiration': '', 'notes': ''},
                {'desc': 'Paper Analyz Medtoxscan Therma', 'material': '', 'supplier': '10333259', 'hand_count': '', 'location': '', 'expiration': '', 'notes': ''},
                {'desc': 'PIPETTE MINIPET DEVICES PV ANA', 'material': '', 'supplier': '10333263', 'hand_count': '', 'location': '', 'expiration': '', 'notes': ''},
                {'desc': 'Hematek Slide Stainer Cleaner', 'material': '', 'supplier': '', 'hand_count': '', 'location': 'AUC', 'expiration': '', 'notes': 'QC slide daily, record in maintenance log'},
            ]
        }
    }
    
    # Standard columns
    columns = [
        {'name': 'DESCRIPTION', 'width': 45},
        {'name': 'MFR#/CAT#', 'width': 20},
        {'name': 'MATERIAL# (KAISER#/OLID)', 'width': 25},
        {'name': 'SUPPLIER ID', 'width': 15},
        {'name': 'ONELINK NUMBER', 'width': 20},
        {'name': 'PAR LEVEL', 'width': 12},
        {'name': 'HAND COUNT', 'width': 12},
        {'name': 'REQ QTY', 'width': 10},
        {'name': 'STATUS', 'width': 15},
        {'name': 'EXPIRATION DATE', 'width': 15},
        {'name': 'LOT NUMBER', 'width': 15},
        {'name': 'LOCATION (MOB/AUC)', 'width': 18},
        {'name': 'LAST UPDATED', 'width': 15},
        {'name': 'UPDATED BY', 'width': 20},
        {'name': 'NOTES', 'width': 40},
        {'name': 'ACTION REQUIRED', 'width': 30}
    ]
    
    # Create sheets for each category
    for category, data in categories.items():
        sheet = wb.create_sheet(title=category)
        
        # Add headers
        for col_idx, col_info in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col_idx, value=col_info['name'])
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            sheet.column_dimensions[get_column_letter(col_idx)].width = col_info['width']
        
        # Add sample data
        for row_idx, item in enumerate(data['items'], 2):
            sheet.cell(row=row_idx, column=1, value=item['desc'])
            sheet.cell(row=row_idx, column=3, value=item['material'])
            sheet.cell(row=row_idx, column=4, value=item['supplier'])
            sheet.cell(row=row_idx, column=7, value=item['hand_count'])
            sheet.cell(row=row_idx, column=10, value=item['expiration'])
            sheet.cell(row=row_idx, column=12, value=item['location'])
            sheet.cell(row=row_idx, column=13, value=datetime.now().strftime('%Y-%m-%d'))
            sheet.cell(row=row_idx, column=14, value='System')
            sheet.cell(row=row_idx, column=15, value=item['notes'])
            
            # Calculate status
            if item['hand_count']:
                try:
                    count = int(item['hand_count'])
                    if count == 0:
                        status = 'OUT OF STOCK'
                        status_color = 'FF0000'
                    elif count < 10:
                        status = 'LOW STOCK'
                        status_color = 'FFA500'
                    else:
                        status = 'OK'
                        status_color = '00FF00'
                    
                    sheet.cell(row=row_idx, column=9, value=status)
                    sheet.cell(row=row_idx, column=9).fill = PatternFill(
                        start_color=status_color, end_color=status_color, fill_type="solid"
                    )
                except:
                    pass
            
            # Highlight expiring items
            if item['expiration'] == '2025-10-31':
                sheet.cell(row=row_idx, column=16, value='REDISTRIBUTE BEFORE EXPIRATION')
                for col in range(1, 17):
                    sheet.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFE5E5", end_color="FFE5E5", fill_type="solid"
                    )
            
            # Add borders
            for col in range(1, 17):
                sheet.cell(row=row_idx, column=col).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Freeze top row
        sheet.freeze_panes = 'A2'
        
        # Add data validation for location
        location_dv = DataValidation(
            type="list",
            formula1='"MOB,AUC,BOTH,MOB→AUC,AUC→MOB"',
            allow_blank=True
        )
        location_dv.add(f'L2:L100')
        sheet.add_data_validation(location_dv)
    
    # Create Summary Dashboard
    summary = wb.create_sheet(title='SUMMARY', index=0)
    
    # Title
    summary.merge_cells('A1:H1')
    title_cell = summary['A1']
    title_cell.value = "LARGO LAB INVENTORY MANAGEMENT SYSTEM"
    title_cell.font = Font(bold=True, size=18, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[1].height = 30
    
    # Lab locations
    summary.merge_cells('A3:D3')
    summary['A3'] = "LAB LOCATIONS"
    summary['A3'].font = Font(bold=True, size=14)
    summary['A4'] = "MOB Lab (Core Lab)"
    summary['A5'] = "AUC Lab (STAT Lab)"
    
    # Critical alerts based on chat
    summary.merge_cells('A7:H7')
    summary['A7'] = "⚠️ CRITICAL ALERTS & ISSUES FROM TECH MEETING"
    summary['A7'].font = Font(bold=True, size=14, color="FF0000")
    
    alerts = [
        "🔴 ALT reagent packs (25 total) expire end of October - need redistribution plan",
        "🟡 Multiple supplier ID numbers found incorrect - needs verification",
        "🟡 MEDTOX QC not logged in Cerner - check maintenance logs",
        "✅ Pneumatic tube system now functional",
        "✅ CBC's can now be run at MOB Lab",
        "🔵 Urine specimens: Primary processor - Lorraine, Backup - Mimi (coordinate with Ingrid)"
    ]
    
    for idx, alert in enumerate(alerts, 9):
        summary[f'A{idx}'] = alert
        summary.merge_cells(f'A{idx}:H{idx}')
    
    # Quick reference
    row = 16
    summary.merge_cells(f'A{row}:H{row}')
    summary[f'A{row}'] = "📋 QUICK REFERENCE - WHO DOES WHAT"
    summary[f'A{row}'].font = Font(bold=True, size=14)
    
    references = [
        "• Inventory Lists Management: Lorraine (Primary), Ingrid (Secondary)",
        "• Supply Ordering: Check with Nathaniel Burmeister for max order quantities",
        "• Urine Processing: Lorraine (Primary), Mimi (Backup when Lorraine is out)",
        "• QC Maintenance Logs: Check green binders - Chemistry, Hematology, Coagulation",
        "• Overflow Testing: Send to AUC when MOB is overwhelmed"
    ]
    
    for idx, ref in enumerate(references, row + 2):
        summary[f'A{idx}'] = ref
        summary.merge_cells(f'A{idx}:H{idx}')
    
    # Instructions
    row = row + 8
    summary.merge_cells(f'A{row}:H{row}')
    summary[f'A{row}'] = "📖 INSTRUCTIONS FOR USE"
    summary[f'A{row}'].font = Font(bold=True, size=14)
    
    instructions = [
        "1. DAILY TASKS:",
        "   • Check hand counts at your assigned bench",
        "   • Update counts in this spreadsheet",
        "   • Note any discrepancies in supplier IDs",
        "",
        "2. WHEN RECEIVING SUPPLIES:",
        "   • Verify supplier ID matches packing slip",
        "   • Update hand count and expiration date",
        "   • Check lot numbers for critical items",
        "",
        "3. FOR LOW STOCK ITEMS (<10):",
        "   • Enter requested quantity in REQ QTY column",
        "   • Notify Nathaniel for ordering",
        "   • Consider redistribution between MOB/AUC",
        "",
        "4. IMPORTANT REMINDERS:",
        "   • Clock in 30 min early for inventory with manager approval",
        "   • Submit Q4 PTO by deadline - seniority rules apply",
        "   • Save and upload to Teams/SharePoint after updates"
    ]
    
    for idx, instruction in enumerate(instructions, row + 2):
        summary[f'A{idx}'] = instruction
        if instruction and not instruction.startswith('   '):
            summary[f'A{idx}'].font = Font(bold=True)
    
    # Adjust column widths for summary
    for col in range(1, 9):
        summary.column_dimensions[get_column_letter(col)].width = 15
    
    # Create Expiring Items sheet
    expiring = wb.create_sheet(title='EXPIRING_ITEMS')
    
    # Headers
    exp_headers = ['DESCRIPTION', 'LOCATION', 'QUANTITY', 'EXPIRATION DATE', 'DAYS UNTIL EXPIRY', 'ACTION PLAN']
    for col_idx, header in enumerate(exp_headers, 1):
        cell = expiring.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add ALT reagent info
    expiring['A2'] = 'ALT Reagent Packs'
    expiring['B2'] = 'MOB/AUC'
    expiring['C2'] = '25'
    expiring['D2'] = '2025-10-31'
    expiring['E2'] = '=D2-TODAY()'
    expiring['F2'] = 'Contact other Kaiser locations for redistribution'
    
    # Adjust widths
    expiring.column_dimensions['A'].width = 30
    expiring.column_dimensions['B'].width = 15
    expiring.column_dimensions['C'].width = 12
    expiring.column_dimensions['D'].width = 20
    expiring.column_dimensions['E'].width = 20
    expiring.column_dimensions['F'].width = 40
    
    return wb

def main():
    print("Creating Comprehensive Lab Inventory Template")
    print("=" * 60)
    
    # Create the template
    wb = create_inventory_template()
    
    # Save the file
    downloads_path = Path("/Users/ugochi141/Downloads")
    filename = "LARGO_LAB_INVENTORY_TEMPLATE_COMPLETE.xlsx"
    filepath = downloads_path / filename
    wb.save(filepath)
    
    # Also save to project
    project_path = Path("/Users/ugochi141/Desktop/LabAutomation/data")
    project_path.mkdir(exist_ok=True)
    project_file = project_path / filename
    wb.save(project_file)
    
    print(f"✅ Template created: {filename}")
    print(f"   Downloads: {filepath}")
    print(f"   Project: {project_file}")
    
    print("\n" + "=" * 60)
    print("📋 TEMPLATE FEATURES:")
    print("=" * 60)
    print("✓ Pre-populated with items from tech meeting discussion")
    print("✓ ALT reagent expiration alert included")
    print("✓ Supplier ID fields ready for correction")
    print("✓ Location dropdown (MOB/AUC/BOTH)")
    print("✓ Status auto-calculation based on hand count")
    print("✓ Expiring items tracking sheet")
    print("✓ Staff assignments and procedures")
    print("✓ Critical alerts from recent issues")
    
    print("\n🎯 NEXT STEPS:")
    print("1. Open the template and review all tabs")
    print("2. Have techs fill in missing hand counts")
    print("3. Verify and correct supplier IDs")
    print("4. Share via Teams with edit permissions")
    print("5. Schedule training with Lorraine and Ingrid")

if __name__ == "__main__":
    main()



