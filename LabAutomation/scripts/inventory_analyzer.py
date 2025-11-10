#!/usr/bin/env python3
"""
Analyze the structure of inventory Excel files to understand data format
"""

import openpyxl
from pathlib import Path

def analyze_excel_file(filepath):
    """Analyze the structure of an Excel file"""
    print(f"\nAnalyzing: {filepath.name}")
    print("=" * 60)
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\nSheet: {sheet_name}")
            print(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
            
            if sheet.max_row > 0:
                # Print first 5 rows to understand structure
                print("\nFirst 5 rows:")
                for row_idx in range(1, min(6, sheet.max_row + 1)):
                    row_data = []
                    for col_idx in range(1, min(sheet.max_column + 1, 10)):  # First 10 columns
                        cell_value = sheet.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            row_data.append(str(cell_value)[:30])  # Truncate long values
                    if row_data:
                        print(f"Row {row_idx}: {row_data}")
                
                # Check for specific columns
                if sheet.max_row > 0:
                    headers = []
                    for col in range(1, sheet.max_column + 1):
                        header = sheet.cell(row=1, column=col).value
                        if header:
                            headers.append(header)
                    if headers:
                        print(f"\nHeaders found: {headers}")
    except Exception as e:
        print(f"Error analyzing file: {e}")

# Analyze key inventory files
downloads_path = Path("/Users/ugochi141/Downloads")

# Main file with hand counts
main_file = downloads_path / "Inventory List Largo 03MAR25_Shanthi and Nate B_09092025.xlsx"
if main_file.exists():
    analyze_excel_file(main_file)

# Check a chemistry-specific file
chem_file = downloads_path / "Inventory List Largo 03MAR25 chem.xlsx"
if chem_file.exists():
    analyze_excel_file(chem_file)

# Check the general inventory spreadsheet
general_file = downloads_path / "Lab Inventory Spreadsheet.xlsx"
if general_file.exists():
    analyze_excel_file(general_file)



