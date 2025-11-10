#!/usr/bin/env python3
"""
Automated Inventory Email System
Monitors inventory levels and emails supply coordinator when reordering is needed
"""

import openpyxl
from pathlib import Path
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import json
import os

class InventoryEmailAutomation:
    def __init__(self, inventory_file, config_file=None):
        self.inventory_file = Path(inventory_file)
        self.items_to_order = []
        self.critical_items = []
        self.expiring_items = []
        
        # Load email configuration
        if config_file and Path(config_file).exists():
            with open(config_file, 'r') as f:
                self.config = json.load(f)
        else:
            # Default configuration
            self.config = {
                "email_settings": {
                    "smtp_server": "smtp.kaiser.org",  # Kaiser's SMTP server
                    "smtp_port": 587,
                    "use_tls": True,
                    "sender_email": "lab.inventory@kaiser.org",
                    "sender_name": "Largo Lab Inventory System",
                    "sender_password": ""  # Will need to be configured
                },
                "recipients": {
                    "supply_coordinator": {
                        "name": "Nathaniel Burmeister",
                        "email": "nathaniel.burmeister@kaiser.org"
                    },
                    "cc_recipients": [
                        {"name": "Lorraine", "email": "lorraine@kaiser.org"},
                        {"name": "Ingrid", "email": "ingrid.benitez-ruiz@kaiser.org"}
                    ],
                    "urgent_recipients": [
                        {"name": "John F Ekpe", "email": "john.ekpe@kaiser.org"},
                        {"name": "Maxwell Booker", "email": "maxwell.booker@kaiser.org"}
                    ]
                },
                "thresholds": {
                    "critical_days_remaining": 7,
                    "urgent_percentage": 25  # Order urgently if stock is 25% or less of PAR
                }
            }
    
    def check_inventory_levels(self):
        """Check all items against reorder points"""
        wb = openpyxl.load_workbook(self.inventory_file, data_only=True)
        
        for sheet_name in ['CHEMISTRY', 'HEMATOLOGY', 'URINALYSIS', 'KITS', 'MISCELLANEOUS']:
            if sheet_name not in wb.sheetnames:
                continue
                
            sheet = wb[sheet_name]
            
            # Column indices (1-based)
            cols = {
                'item_num': 1,
                'description': 2,
                'manufacturer': 3,
                'catalog': 4,
                'material': 5,
                'supplier_id': 6,
                'onelink': 7,
                'package_size': 8,
                'unit': 9,
                'par': 10,
                'min_stock': 11,
                'max_stock': 12,
                'hand_count': 13,
                'req_qty': 14,
                'reorder_point': 15,
                'status': 16,
                'expiration': 17,
                'location': 19,
                'notes': 29
            }
            
            # Check each item
            for row in range(2, sheet.max_row + 1):
                # Get item details
                item = {
                    'category': sheet_name,
                    'item_num': sheet.cell(row=row, column=cols['item_num']).value,
                    'description': sheet.cell(row=row, column=cols['description']).value,
                    'manufacturer': sheet.cell(row=row, column=cols['manufacturer']).value,
                    'catalog': sheet.cell(row=row, column=cols['catalog']).value,
                    'material': sheet.cell(row=row, column=cols['material']).value,
                    'supplier_id': sheet.cell(row=row, column=cols['supplier_id']).value,
                    'onelink': sheet.cell(row=row, column=cols['onelink']).value,
                    'package_size': sheet.cell(row=row, column=cols['package_size']).value,
                    'unit': sheet.cell(row=row, column=cols['unit']).value,
                    'par': self._safe_int(sheet.cell(row=row, column=cols['par']).value),
                    'min_stock': self._safe_int(sheet.cell(row=row, column=cols['min_stock']).value),
                    'hand_count': self._safe_int(sheet.cell(row=row, column=cols['hand_count']).value),
                    'reorder_point': self._safe_int(sheet.cell(row=row, column=cols['reorder_point']).value),
                    'location': sheet.cell(row=row, column=cols['location']).value,
                    'notes': sheet.cell(row=row, column=cols['notes']).value
                }
                
                # Skip if no description
                if not item['description']:
                    continue
                
                # Check if needs ordering
                if item['hand_count'] is not None and item['reorder_point'] is not None:
                    if item['hand_count'] <= item['reorder_point']:
                        # Calculate order quantity
                        if item['par'] and item['hand_count'] is not None:
                            order_qty = item['par'] - item['hand_count']
                            # Round up to package size if specified
                            if item['min_stock']:
                                order_qty = max(order_qty, item['min_stock'])
                        else:
                            order_qty = item['reorder_point']
                        
                        item['order_quantity'] = order_qty
                        item['urgency'] = self._calculate_urgency(item)
                        
                        self.items_to_order.append(item)
                        
                        if item['urgency'] == 'CRITICAL':
                            self.critical_items.append(item)
                
                # Check for ALT reagents specifically
                if 'ALT' in str(item['description']).upper() and 'REAGENT' in str(item['description']).upper():
                    if item['hand_count'] and item['hand_count'] > 20:
                        item['excess_quantity'] = item['hand_count'] - (item['par'] or 8)
                        item['expiration_note'] = 'EXPIRES OCT 31 - MUST REDISTRIBUTE!'
                        self.expiring_items.append(item)
    
    def _safe_int(self, value):
        """Safely convert to integer"""
        if value is None or value == '':
            return None
        try:
            return int(value)
        except:
            return None
    
    def _calculate_urgency(self, item):
        """Calculate urgency level"""
        if item['hand_count'] == 0:
            return 'CRITICAL'
        elif item['hand_count'] and item['min_stock'] and item['hand_count'] <= item['min_stock']:
            return 'URGENT'
        else:
            return 'ROUTINE'
    
    def generate_order_email(self):
        """Generate HTML email with order details"""
        if not self.items_to_order and not self.expiring_items:
            return None
            
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                h1 {{ color: #366092; }}
                h2 {{ color: #366092; }}
                .critical {{ background-color: #ffcccc; font-weight: bold; }}
                .urgent {{ background-color: #fff3cd; font-weight: bold; }}
                .routine {{ background-color: #d4edda; }}
                table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #366092; color: white; }}
                .alert {{ background-color: #ff6666; color: white; padding: 10px; margin: 10px 0; }}
                .note {{ background-color: #f0f0f0; padding: 10px; margin: 10px 0; }}
            </style>
        </head>
        <body>
            <h1>Largo Lab Inventory Order Request</h1>
            <p><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
            <p><strong>From:</strong> MOB Lab (Core) & AUC Lab (STAT)</p>
        """
        
        # Add ALT reagent alert if needed
        if self.expiring_items:
            html += """
            <div class="alert">
                <h2>⚠️ URGENT: EXPIRING ITEMS - ACTION REQUIRED</h2>
            """
            for item in self.expiring_items:
                html += f"""
                <p><strong>{item['description']}</strong><br>
                Current Stock: {item['hand_count']} packs<br>
                PAR Level: {item['par'] or 8} packs<br>
                EXCESS TO REDISTRIBUTE: {item['excess_quantity']} packs<br>
                {item['expiration_note']}<br>
                <strong>Contact other Kaiser locations immediately!</strong></p>
                """
            html += "</div>"
        
        # Add critical items section
        if self.critical_items:
            html += """
            <h2>🔴 CRITICAL - OUT OF STOCK ITEMS</h2>
            <table>
                <tr>
                    <th>Item</th>
                    <th>Catalog #</th>
                    <th>Supplier ID</th>
                    <th>Current</th>
                    <th>Order Qty</th>
                    <th>Location</th>
                </tr>
            """
            for item in self.critical_items:
                html += f"""
                <tr class="critical">
                    <td>{item['description']}</td>
                    <td>{item['catalog'] or 'N/A'}</td>
                    <td>{item['supplier_id'] or 'CHECK'}</td>
                    <td>{item['hand_count']}</td>
                    <td>{item['order_quantity']} {item['unit'] or ''}</td>
                    <td>{item['location'] or 'TBD'}</td>
                </tr>
                """
            html += "</table>"
        
        # Add regular order items
        urgent_items = [i for i in self.items_to_order if i['urgency'] == 'URGENT' and i not in self.critical_items]
        routine_items = [i for i in self.items_to_order if i['urgency'] == 'ROUTINE']
        
        if urgent_items:
            html += """
            <h2>🟡 URGENT - LOW STOCK ITEMS</h2>
            <table>
                <tr>
                    <th>Item</th>
                    <th>Manufacturer</th>
                    <th>Catalog #</th>
                    <th>Supplier ID</th>
                    <th>Package Size</th>
                    <th>Current</th>
                    <th>Reorder Point</th>
                    <th>Order Qty</th>
                    <th>Location</th>
                </tr>
            """
            for item in urgent_items:
                html += f"""
                <tr class="urgent">
                    <td>{item['description']}</td>
                    <td>{item['manufacturer'] or ''}</td>
                    <td>{item['catalog'] or ''}</td>
                    <td>{item['supplier_id'] or 'VERIFY'}</td>
                    <td>{item['package_size'] or ''}</td>
                    <td>{item['hand_count']}</td>
                    <td>{item['reorder_point']}</td>
                    <td>{item['order_quantity']} {item['unit'] or ''}</td>
                    <td>{item['location'] or 'TBD'}</td>
                </tr>
                """
            html += "</table>"
        
        if routine_items:
            html += """
            <h2>🟢 ROUTINE - STANDARD REORDER</h2>
            <table>
                <tr>
                    <th>Item</th>
                    <th>Catalog #</th>
                    <th>Supplier ID</th>
                    <th>Current</th>
                    <th>PAR</th>
                    <th>Order Qty</th>
                </tr>
            """
            for item in routine_items:
                html += f"""
                <tr class="routine">
                    <td>{item['description']}</td>
                    <td>{item['catalog'] or ''}</td>
                    <td>{item['supplier_id'] or ''}</td>
                    <td>{item['hand_count']}</td>
                    <td>{item['par']}</td>
                    <td>{item['order_quantity']} {item['unit'] or ''}</td>
                </tr>
                """
            html += "</table>"
        
        # Add notes section
        html += """
        <div class="note">
            <h3>Important Notes:</h3>
            <ul>
                <li><strong>Supplier ID Verification:</strong> Several supplier IDs need verification (marked as VERIFY or CHECK)</li>
                <li><strong>MEDTOX QC:</strong> Remember to log all QC in Cerner when received</li>
                <li><strong>Delivery Location:</strong> Split deliveries between MOB and AUC as indicated</li>
                <li><strong>Questions:</strong> Contact Lorraine (primary) or Ingrid (backup)</li>
            </ul>
        </div>
        
        <p><strong>Lab Contacts:</strong><br>
        MOB Lab: Lorraine (Primary)<br>
        AUC Lab: Ingrid (Secondary)<br>
        Tech Lead: Maxwell Booker (Supplier ID verification)</p>
        
        <p><em>This is an automated message from the Largo Lab Inventory Management System</em></p>
        </body>
        </html>
        """
        
        return html
    
    def send_order_email(self, test_mode=True):
        """Send the order email"""
        html_content = self.generate_order_email()
        
        if not html_content:
            print("No items need ordering at this time.")
            return
        
        # Create message
        msg = MIMEMultipart('mixed')
        msg['Subject'] = f"Lab Supply Order Request - {datetime.now().strftime('%Y-%m-%d')}"
        
        # Add urgent marker if critical items
        if self.critical_items or self.expiring_items:
            msg['Subject'] = "URGENT: " + msg['Subject']
        
        msg['From'] = f"{self.config['email_settings']['sender_name']} <{self.config['email_settings']['sender_email']}>"
        msg['To'] = self.config['recipients']['supply_coordinator']['email']
        
        # Add CC recipients
        cc_emails = [r['email'] for r in self.config['recipients']['cc_recipients']]
        if self.critical_items:
            # Add urgent recipients for critical items
            cc_emails.extend([r['email'] for r in self.config['recipients']['urgent_recipients']])
        msg['Cc'] = ', '.join(cc_emails)
        
        # Attach HTML content
        msg.attach(MIMEText(html_content, 'html'))
        
        # Attach Excel file
        with open(self.inventory_file, 'rb') as f:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename="Lab_Inventory_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            )
            msg.attach(part)
        
        if test_mode:
            # Save email to file instead of sending
            test_file = Path("test_order_email.html")
            with open(test_file, 'w') as f:
                f.write(html_content)
            print(f"Test email saved to: {test_file}")
            
            # Also save email details
            email_details = {
                "to": self.config['recipients']['supply_coordinator']['email'],
                "cc": cc_emails,
                "subject": msg['Subject'],
                "items_to_order": len(self.items_to_order),
                "critical_items": len(self.critical_items),
                "expiring_items": len(self.expiring_items)
            }
            
            with open("email_details.json", 'w') as f:
                json.dump(email_details, f, indent=2)
            
            print("\nEmail Details:")
            print(f"To: {email_details['to']}")
            print(f"CC: {', '.join(email_details['cc'])}")
            print(f"Subject: {email_details['subject']}")
            print(f"Items to order: {email_details['items_to_order']}")
            print(f"Critical items: {email_details['critical_items']}")
            print(f"Expiring items: {email_details['expiring_items']}")
        else:
            # Send actual email
            try:
                server = smtplib.SMTP(
                    self.config['email_settings']['smtp_server'],
                    self.config['email_settings']['smtp_port']
                )
                if self.config['email_settings']['use_tls']:
                    server.starttls()
                
                # Login if password provided
                if self.config['email_settings']['sender_password']:
                    server.login(
                        self.config['email_settings']['sender_email'],
                        self.config['email_settings']['sender_password']
                    )
                
                # Send email
                all_recipients = [msg['To']] + cc_emails
                server.send_message(msg, to_addrs=all_recipients)
                server.quit()
                
                print("✅ Order email sent successfully!")
                print(f"To: {msg['To']}")
                print(f"CC: {msg['Cc']}")
                
                # Log the order
                self.log_order()
                
            except Exception as e:
                print(f"❌ Error sending email: {e}")
                print("Email content saved to test_order_email.html")
                with open("test_order_email.html", 'w') as f:
                    f.write(html_content)
    
    def log_order(self):
        """Log the order for tracking"""
        log_entry = {
            "timestamp": datetime.now().isoformat(),
            "items_ordered": len(self.items_to_order),
            "critical_items": len(self.critical_items),
            "total_value": "TBD",
            "order_details": []
        }
        
        for item in self.items_to_order:
            log_entry["order_details"].append({
                "item": item['description'],
                "quantity": item['order_quantity'],
                "urgency": item['urgency']
            })
        
        # Append to log file
        log_file = Path("order_history.json")
        if log_file.exists():
            with open(log_file, 'r') as f:
                history = json.load(f)
        else:
            history = []
        
        history.append(log_entry)
        
        with open(log_file, 'w') as f:
            json.dump(history, f, indent=2)

def create_email_config():
    """Create email configuration file"""
    config = {
        "email_settings": {
            "smtp_server": "smtp.kaiser.org",
            "smtp_port": 587,
            "use_tls": True,
            "sender_email": "lab.inventory@kaiser.org",
            "sender_name": "Largo Lab Inventory System",
            "sender_password": ""  # Must be set securely
        },
        "recipients": {
            "supply_coordinator": {
                "name": "Nathaniel Burmeister",
                "email": "nathaniel.burmeister@kaiser.org"
            },
            "cc_recipients": [
                {"name": "Lorraine", "email": "lorraine@kaiser.org"},
                {"name": "Ingrid Benitez-Ruiz", "email": "ingrid.benitez-ruiz@kaiser.org"}
            ],
            "urgent_recipients": [
                {"name": "John F Ekpe", "email": "john.ekpe@kaiser.org"},
                {"name": "Maxwell Booker", "email": "maxwell.booker@kaiser.org"}
            ]
        },
        "schedule": {
            "daily_check_time": "07:00",
            "weekly_summary_day": "Monday",
            "auto_send": False  # Set to True when ready for production
        }
    }
    
    config_path = Path("/Users/ugochi141/Desktop/LabAutomation/config/inventory/email_config.json")
    config_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(config_path, 'w') as f:
        json.dump(config, f, indent=2)
    
    print(f"Email configuration saved to: {config_path}")
    return config_path

def main():
    print("Lab Inventory Email Automation System")
    print("=" * 60)
    
    # Create configuration if needed
    config_path = Path("/Users/ugochi141/Desktop/LabAutomation/config/inventory/email_config.json")
    if not config_path.exists():
        config_path = create_email_config()
    
    # Find latest inventory file
    inventory_path = Path("/Users/ugochi141/Downloads/LARGO_LAB_INVENTORY_WITH_PAR_LEVELS.xlsx")
    if not inventory_path.exists():
        inventory_path = Path("/Users/ugochi141/Desktop/LabAutomation/data/inventory/LARGO_LAB_INVENTORY_WITH_PAR_LEVELS.xlsx")
    
    if not inventory_path.exists():
        print("❌ Inventory file not found!")
        return
    
    print(f"Using inventory file: {inventory_path.name}")
    
    # Initialize automation
    automation = InventoryEmailAutomation(inventory_path, config_path)
    
    # Check inventory levels
    print("\nChecking inventory levels...")
    automation.check_inventory_levels()
    
    print(f"\nItems needing reorder: {len(automation.items_to_order)}")
    print(f"Critical items: {len(automation.critical_items)}")
    print(f"Expiring items: {len(automation.expiring_items)}")
    
    # Send order email (test mode)
    if automation.items_to_order or automation.expiring_items:
        print("\nGenerating order email...")
        automation.send_order_email(test_mode=True)
        
        print("\n" + "=" * 60)
        print("✅ Email automation complete!")
        print("📧 Test email saved to: test_order_email.html")
        print("📊 Order details saved to: email_details.json")
        print("\nTo send actual emails:")
        print("1. Update email_config.json with SMTP credentials")
        print("2. Set 'auto_send' to True")
        print("3. Run with test_mode=False")
    else:
        print("\n✅ All inventory levels are adequate - no orders needed")

if __name__ == "__main__":
    main()



