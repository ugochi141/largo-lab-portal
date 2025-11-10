#!/usr/bin/env python3
"""
Create a DETAILED inventory template with comprehensive information for all supplies
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
from pathlib import Path

def create_detailed_inventory():
    """Create the most detailed inventory with all specifications"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Extended columns for detailed tracking
    columns = [
        {'name': 'ITEM #', 'width': 10},
        {'name': 'DESCRIPTION', 'width': 45},
        {'name': 'MANUFACTURER', 'width': 20},
        {'name': 'MFR#/CAT#', 'width': 20},
        {'name': 'MATERIAL# (KAISER#/OLID)', 'width': 25},
        {'name': 'SUPPLIER ID', 'width': 15},
        {'name': 'ONELINK NUMBER', 'width': 20},
        {'name': 'PACKAGE SIZE', 'width': 15},
        {'name': 'UNIT OF MEASURE', 'width': 12},
        {'name': 'PAR LEVEL', 'width': 12},
        {'name': 'MIN STOCK', 'width': 12},
        {'name': 'MAX STOCK', 'width': 12},
        {'name': 'HAND COUNT', 'width': 12},
        {'name': 'REQ QTY', 'width': 10},
        {'name': 'REORDER POINT', 'width': 15},
        {'name': 'STATUS', 'width': 15},
        {'name': 'EXPIRATION DATE', 'width': 15},
        {'name': 'LOT NUMBER', 'width': 15},
        {'name': 'LOCATION (MOB/AUC)', 'width': 18},
        {'name': 'STORAGE TEMP', 'width': 15},
        {'name': 'STORAGE LOCATION', 'width': 20},
        {'name': 'ANALYZER/EQUIPMENT', 'width': 20},
        {'name': 'TEST/PROCEDURE', 'width': 25},
        {'name': 'CRITICAL ITEM', 'width': 12},
        {'name': 'LAST ORDERED', 'width': 15},
        {'name': 'LAST RECEIVED', 'width': 15},
        {'name': 'LAST UPDATED', 'width': 15},
        {'name': 'UPDATED BY', 'width': 20},
        {'name': 'NOTES', 'width': 40},
        {'name': 'ACTION REQUIRED', 'width': 30}
    ]
    
    # CHEMISTRY - Detailed items
    chemistry_items = [
        # Critical ALT Reagents
        {
            'item_num': 'CH001',
            'desc': 'ALT (Alanine Aminotransferase) Reagent Pack',
            'manufacturer': 'Roche Diagnostics',
            'mfr_cat': '07414463190',
            'material': '07414463190',
            'supplier': '10333255',
            'package_size': '800 tests',
            'unit': 'Pack',
            'par': '5',
            'min': '2',
            'max': '10',
            'hand_count': '25',
            'storage_temp': '2-8°C',
            'storage_loc': 'Refrigerator #1',
            'analyzer': 'Roche Cobas c303/c503',
            'test': 'ALT/SGPT',
            'critical': 'YES',
            'notes': '⚠️ 25 PACKS EXPIRE OCT 31 - MUST REDISTRIBUTE TO OTHER KAISER LOCATIONS',
            'action': 'URGENT - Contact other labs TODAY'
        },
        # Other Chemistry Reagents
        {
            'item_num': 'CH002',
            'desc': 'ASTP (Aspartate Aminotransferase) Reagent',
            'manufacturer': 'Roche Diagnostics',
            'mfr_cat': '07876866190',
            'material': '07876866190',
            'supplier': '',
            'package_size': '800 tests',
            'unit': 'Pack',
            'par': '5',
            'min': '2',
            'max': '10',
            'storage_temp': '2-8°C',
            'storage_loc': 'Refrigerator #1',
            'analyzer': 'Roche Cobas c303/c503',
            'test': 'AST/SGOT',
            'critical': 'YES'
        },
        {
            'item_num': 'CH003',
            'desc': 'Glucose (GLUC3) Reagent',
            'manufacturer': 'Roche Diagnostics',
            'mfr_cat': '04404483190',
            'material': '04404483190',
            'supplier': '',
            'package_size': '1000 tests',
            'unit': 'Pack',
            'par': '10',
            'min': '5',
            'max': '20',
            'storage_temp': '2-8°C',
            'storage_loc': 'Refrigerator #1',
            'analyzer': 'Roche Cobas c303/c503',
            'test': 'Glucose',
            'critical': 'YES',
            'notes': 'High volume test'
        },
        {
            'item_num': 'CH004',
            'desc': 'Creatinine (CREJ2) Enzymatic Reagent',
            'manufacturer': 'Roche Diagnostics',
            'mfr_cat': '06407137190',
            'material': '06407137190',
            'supplier': '',
            'package_size': '1200 tests',
            'unit': 'Pack',
            'par': '8',
            'min': '4',
            'max': '15',
            'storage_temp': '2-8°C',
            'storage_loc': 'Refrigerator #1',
            'analyzer': 'Roche Cobas c303/c503',
            'test': 'Creatinine',
            'critical': 'YES',
            'notes': 'Essential for renal panels'
        },
        {
            'item_num': 'CH005',
            'desc': 'Total Bilirubin (BILT3) Reagent',
            'manufacturer': 'Roche Diagnostics',
            'mfr_cat': '05795419190',
            'material': '05795419190',
            'supplier': '',
            'package_size': '600 tests',
            'unit': 'Pack',
            'par': '3',
            'min': '1',
            'max': '6',
            'storage_temp': '2-8°C',
            'storage_loc': 'Refrigerator #1',
            'analyzer': 'Roche Cobas c303/c503',
            'test': 'Total Bilirubin',
            'critical': 'YES'
        },
        {
            'item_num': 'CH006',
            'desc': 'Direct Bilirubin (D BILI) Reagent',
            'manufacturer': 'Roche Diagnostics',
            'mfr_cat': '07191138190',
            'material': '07191138190',
            'supplier': '',
            'package_size': '300 tests',
            'unit': 'Pack',
            'par': '2',
            'min': '1',
            'max': '4',
            'storage_temp': '2-8°C',
            'storage_loc': 'Refrigerator #1',
            'analyzer': 'Roche Cobas c303/c503',
            'test': 'Direct Bilirubin',
            'critical': 'NO'
        },
        {
            'item_num': 'CH007',
            'desc': 'Alkaline Phosphatase (ALP) IFCC Gen.2',
            'manufacturer': 'Roche Diagnostics',
            'mfr_cat': '03314839190',
            'material': '03314839190',
            'supplier': '',
            'package_size': '800 tests',
            'unit': 'Pack',
            'par': '3',
            'min': '1',
            'max': '6',
            'storage_temp': '2-8°C',
            'storage_loc': 'Refrigerator #1',
            'analyzer': 'Roche Cobas c303/c503',
            'test': 'Alkaline Phosphatase',
            'critical': 'YES'
        },
        {
            'item_num': 'CH008',
            'desc': 'Albumin (ALB2) BCG Method',
            'manufacturer': 'Roche Diagnostics',
            'mfr_cat': '07511914190',
            'material': '07511914190',
            'supplier': '',
            'package_size': '600 tests',
            'unit': 'Pack',
            'par': '4',
            'min': '2',
            'max': '8',
            'storage_temp': '2-8°C',
            'storage_loc': 'Refrigerator #1',
            'analyzer': 'Roche Cobas c303/c503',
            'test': 'Albumin',
            'critical': 'YES'
        },
        {
            'item_num': 'CH009',
            'desc': 'Total Protein (TP2) Reagent',
            'manufacturer': 'Roche Diagnostics',
            'mfr_cat': '03183734190',
            'material': '03183734190',
            'supplier': '',
            'package_size': '600 tests',
            'unit': 'Pack',
            'par': '3',
            'min': '1',
            'max': '6',
            'storage_temp': '2-8°C',
            'storage_loc': 'Refrigerator #1',
            'analyzer': 'Roche Cobas c303/c503',
            'test': 'Total Protein',
            'critical': 'NO'
        },
        {
            'item_num': 'CH010',
            'desc': 'Calcium (CA2) Reagent',
            'manufacturer': 'Roche Diagnostics',
            'mfr_cat': '05061431190',
            'material': '05061431190',
            'supplier': '',
            'package_size': '600 tests',
            'unit': 'Pack',
            'par': '4',
            'min': '2',
            'max': '8',
            'storage_temp': '2-8°C',
            'storage_loc': 'Refrigerator #1',
            'analyzer': 'Roche Cobas c303/c503',
            'test': 'Calcium',
            'critical': 'YES'
        },
        # ISE Consumables
        {
            'item_num': 'CH020',
            'desc': 'ISE Reference Electrode Gen 2',
            'manufacturer': 'Roche Diagnostics',
            'mfr_cat': '10753408001',
            'material': '10753408001',
            'supplier': '',
            'package_size': '1 unit',
            'unit': 'Each',
            'par': '2',
            'min': '1',
            'max': '3',
            'storage_temp': 'Room Temp',
            'storage_loc': 'Supply Room A',
            'analyzer': 'Roche ISE Module',
            'test': 'Electrolytes',
            'critical': 'YES',
            'notes': 'Replace every 4-6 weeks'
        },
        {
            'item_num': 'CH021',
            'desc': 'ISE Chloride Electrode Cartridge',
            'manufacturer': 'Roche Diagnostics',
            'mfr_cat': '3246353001',
            'material': '3246353001',
            'supplier': '',
            'package_size': '1 unit',
            'unit': 'Each',
            'par': '2',
            'min': '1',
            'max': '3',
            'storage_temp': 'Room Temp',
            'storage_loc': 'Supply Room A',
            'analyzer': 'Roche ISE Module',
            'test': 'Chloride',
            'critical': 'YES'
        },
        {
            'item_num': 'CH022',
            'desc': 'ISE Potassium Electrode Cartridge',
            'manufacturer': 'Roche Diagnostics',
            'mfr_cat': '10825441001',
            'material': '10825441001',
            'supplier': '',
            'package_size': '1 unit',
            'unit': 'Each',
            'par': '2',
            'min': '1',
            'max': '3',
            'storage_temp': 'Room Temp',
            'storage_loc': 'Supply Room A',
            'analyzer': 'Roche ISE Module',
            'test': 'Potassium',
            'critical': 'YES'
        },
        {
            'item_num': 'CH023',
            'desc': 'ISE Sodium Electrode Cartridge Gen 2',
            'manufacturer': 'Roche Diagnostics',
            'mfr_cat': '10753416001',
            'material': '10753416001',
            'supplier': '',
            'package_size': '1 unit',
            'unit': 'Each',
            'par': '2',
            'min': '1',
            'max': '3',
            'storage_temp': 'Room Temp',
            'storage_loc': 'Supply Room A',
            'analyzer': 'Roche ISE Module',
            'test': 'Sodium',
            'critical': 'YES'
        },
        # Maintenance Solutions
        {
            'item_num': 'CH030',
            'desc': 'Acid Wash III Solution',
            'manufacturer': 'Roche Diagnostics',
            'mfr_cat': '10731897001',
            'material': '10731897001',
            'supplier': '',
            'package_size': '2L canister',
            'unit': 'Canister',
            'par': '4',
            'min': '2',
            'max': '8',
            'storage_temp': 'Room Temp',
            'storage_loc': 'Supply Room B',
            'analyzer': 'Roche Cobas',
            'test': 'Maintenance',
            'critical': 'NO',
            'notes': 'Weekly maintenance'
        },
        {
            'item_num': 'CH031',
            'desc': 'Basic Wash Solution',
            'manufacturer': 'Roche Diagnostics',
            'mfr_cat': '10931805001',
            'material': '10931805001',
            'supplier': '',
            'package_size': '2L canister',
            'unit': 'Canister',
            'par': '4',
            'min': '2',
            'max': '8',
            'storage_temp': 'Room Temp',
            'storage_loc': 'Supply Room B',
            'analyzer': 'Roche Cobas',
            'test': 'Maintenance',
            'critical': 'NO'
        }
    ]
    
    # HEMATOLOGY - Detailed items
    hematology_items = [
        {
            'item_num': 'HE001',
            'desc': 'CBC Control - Level 1 (Low)',
            'manufacturer': 'Sysmex',
            'mfr_cat': 'XN-CHECK-L1',
            'material': '',
            'supplier': '',
            'package_size': '3mL x 6 vials',
            'unit': 'Box',
            'par': '2',
            'min': '1',
            'max': '4',
            'storage_temp': '2-8°C',
            'storage_loc': 'Hematology Fridge',
            'analyzer': 'Sysmex XN-1000',
            'test': 'CBC QC',
            'critical': 'YES',
            'notes': 'Run daily at start of shift'
        },
        {
            'item_num': 'HE002',
            'desc': 'CBC Control - Level 2 (Normal)',
            'manufacturer': 'Sysmex',
            'mfr_cat': 'XN-CHECK-L2',
            'material': '',
            'supplier': '',
            'package_size': '3mL x 6 vials',
            'unit': 'Box',
            'par': '2',
            'min': '1',
            'max': '4',
            'storage_temp': '2-8°C',
            'storage_loc': 'Hematology Fridge',
            'analyzer': 'Sysmex XN-1000',
            'test': 'CBC QC',
            'critical': 'YES',
            'notes': 'Run daily at start of shift'
        },
        {
            'item_num': 'HE003',
            'desc': 'CBC Control - Level 3 (High)',
            'manufacturer': 'Sysmex',
            'mfr_cat': 'XN-CHECK-L3',
            'material': '',
            'supplier': '',
            'package_size': '3mL x 6 vials',
            'unit': 'Box',
            'par': '2',
            'min': '1',
            'max': '4',
            'storage_temp': '2-8°C',
            'storage_loc': 'Hematology Fridge',
            'analyzer': 'Sysmex XN-1000',
            'test': 'CBC QC',
            'critical': 'YES',
            'notes': 'Run daily at start of shift'
        },
        {
            'item_num': 'HE004',
            'desc': 'Cellpack DCL Diluent',
            'manufacturer': 'Sysmex',
            'mfr_cat': 'DCL-20L',
            'material': '',
            'supplier': '',
            'package_size': '20L cubitainer',
            'unit': 'Each',
            'par': '8',
            'min': '4',
            'max': '12',
            'storage_temp': 'Room Temp',
            'storage_loc': 'Hematology Storage',
            'analyzer': 'Sysmex XN-1000',
            'test': 'CBC',
            'critical': 'YES',
            'notes': 'Primary diluent'
        },
        {
            'item_num': 'HE005',
            'desc': 'Stromatolyser 4DL',
            'manufacturer': 'Sysmex',
            'mfr_cat': 'SLS-5L',
            'material': '',
            'supplier': '',
            'package_size': '5L bottle',
            'unit': 'Each',
            'par': '4',
            'min': '2',
            'max': '6',
            'storage_temp': 'Room Temp',
            'storage_loc': 'Hematology Storage',
            'analyzer': 'Sysmex XN-1000',
            'test': 'WBC Differential',
            'critical': 'YES',
            'notes': 'Lyse reagent'
        },
        {
            'item_num': 'HE006',
            'desc': 'Sulfolyser',
            'manufacturer': 'Sysmex',
            'mfr_cat': 'SUL-1L',
            'material': '',
            'supplier': '',
            'package_size': '1L bottle',
            'unit': 'Each',
            'par': '2',
            'min': '1',
            'max': '4',
            'storage_temp': 'Room Temp',
            'storage_loc': 'Hematology Storage',
            'analyzer': 'Sysmex XN-1000',
            'test': 'RBC/PLT',
            'critical': 'YES'
        },
        {
            'item_num': 'HE007',
            'desc': 'Cellclean',
            'manufacturer': 'Sysmex',
            'mfr_cat': 'CLC-5L',
            'material': '',
            'supplier': '',
            'package_size': '5L bottle',
            'unit': 'Each',
            'par': '2',
            'min': '1',
            'max': '4',
            'storage_temp': 'Room Temp',
            'storage_loc': 'Hematology Storage',
            'analyzer': 'Sysmex XN-1000',
            'test': 'Maintenance',
            'critical': 'NO',
            'notes': 'Daily cleaning'
        },
        {
            'item_num': 'HE008',
            'desc': 'Wright-Giemsa Stain',
            'manufacturer': 'Sigma-Aldrich',
            'mfr_cat': 'WG16',
            'material': '',
            'supplier': '',
            'package_size': '500mL bottle',
            'unit': 'Each',
            'par': '2',
            'min': '1',
            'max': '4',
            'storage_temp': 'Room Temp',
            'storage_loc': 'Staining Area',
            'analyzer': 'Hematek Stainer',
            'test': 'Blood Smear',
            'critical': 'YES',
            'notes': 'For differential staining'
        }
    ]
    
    # URINALYSIS - Detailed items
    urinalysis_items = [
        {
            'item_num': 'UA001',
            'desc': 'Urinalysis Test Strips - 10 Parameter',
            'manufacturer': 'Roche',
            'mfr_cat': 'CHEMSTRIP-10',
            'material': '',
            'supplier': '',
            'package_size': '100 strips/bottle',
            'unit': 'Bottle',
            'par': '10',
            'min': '5',
            'max': '20',
            'storage_temp': 'Room Temp',
            'storage_loc': 'UA Bench',
            'analyzer': 'Urisys 1100',
            'test': 'Urinalysis',
            'critical': 'YES',
            'notes': 'Keep tightly closed'
        },
        {
            'item_num': 'UA002',
            'desc': 'UA Control - Normal Level',
            'manufacturer': 'Bio-Rad',
            'mfr_cat': 'QUA-NORM',
            'material': '',
            'supplier': '',
            'package_size': '12mL x 10 vials',
            'unit': 'Box',
            'par': '2',
            'min': '1',
            'max': '4',
            'storage_temp': '2-8°C',
            'storage_loc': 'UA Fridge',
            'analyzer': 'Urisys 1100',
            'test': 'UA QC',
            'critical': 'YES',
            'notes': 'Run each shift'
        },
        {
            'item_num': 'UA003',
            'desc': 'UA Control - Abnormal Level',
            'manufacturer': 'Bio-Rad',
            'mfr_cat': 'QUA-ABN',
            'material': '',
            'supplier': '',
            'package_size': '12mL x 10 vials',
            'unit': 'Box',
            'par': '2',
            'min': '1',
            'max': '4',
            'storage_temp': '2-8°C',
            'storage_loc': 'UA Fridge',
            'analyzer': 'Urisys 1100',
            'test': 'UA QC',
            'critical': 'YES',
            'notes': 'Run each shift'
        },
        {
            'item_num': 'UA004',
            'desc': 'Urine Collection Cups - 120mL',
            'manufacturer': 'Globe Scientific',
            'mfr_cat': '5912',
            'material': '',
            'supplier': '',
            'package_size': '500/case',
            'unit': 'Case',
            'par': '4',
            'min': '2',
            'max': '8',
            'storage_temp': 'Room Temp',
            'storage_loc': 'Supply Room',
            'analyzer': 'N/A',
            'test': 'UA Collection',
            'critical': 'YES',
            'notes': 'Sterile cups'
        },
        {
            'item_num': 'UA005',
            'desc': 'UA Transfer Tubes - 12mL',
            'manufacturer': 'BD',
            'mfr_cat': '364961',
            'material': '',
            'supplier': '',
            'package_size': '100/pack',
            'unit': 'Pack',
            'par': '10',
            'min': '5',
            'max': '20',
            'storage_temp': 'Room Temp',
            'storage_loc': 'UA Bench',
            'analyzer': 'N/A',
            'test': 'UA Processing',
            'critical': 'YES',
            'notes': 'Yellow top, no additive'
        }
    ]
    
    # KITS - Detailed items
    kits_items = [
        {
            'item_num': 'KT001',
            'desc': 'MEDTOX Drug Screen Kit',
            'manufacturer': 'MEDTOX',
            'mfr_cat': 'DT-10',
            'material': '',
            'supplier': '604032',
            'package_size': '25 tests/kit',
            'unit': 'Kit',
            'par': '4',
            'min': '2',
            'max': '8',
            'storage_temp': 'Room Temp',
            'storage_loc': 'Toxicology Area',
            'analyzer': 'MEDTOX Scanner',
            'test': 'Drug Screen',
            'critical': 'YES',
            'notes': 'CHECK QC IS LOGGED IN CERNER!'
        },
        {
            'item_num': 'KT002',
            'desc': 'MEDTOX Positive Control',
            'manufacturer': 'MEDTOX',
            'mfr_cat': 'DT-POS',
            'material': '',
            'supplier': '10333255',
            'package_size': '3mL x 5 vials',
            'unit': 'Box',
            'par': '2',
            'min': '1',
            'max': '4',
            'hand_count': '5',
            'storage_temp': '2-8°C',
            'storage_loc': 'Toxicology Fridge',
            'analyzer': 'MEDTOX Scanner',
            'test': 'Drug Screen QC',
            'critical': 'YES',
            'notes': 'Must log in Cerner'
        },
        {
            'item_num': 'KT003',
            'desc': 'MEDTOX Negative Control',
            'manufacturer': 'MEDTOX',
            'mfr_cat': 'DT-NEG',
            'material': '',
            'supplier': '10283225',
            'package_size': '3mL x 5 vials',
            'unit': 'Box',
            'par': '2',
            'min': '1',
            'max': '4',
            'hand_count': '1',
            'storage_temp': '2-8°C',
            'storage_loc': 'Toxicology Fridge',
            'analyzer': 'MEDTOX Scanner',
            'test': 'Drug Screen QC',
            'critical': 'YES',
            'notes': 'Must log in Cerner'
        },
        {
            'item_num': 'KT004',
            'desc': 'PT/INR Control Set',
            'manufacturer': 'Stago',
            'mfr_cat': 'CK-PREST-2',
            'material': '',
            'supplier': '',
            'package_size': '1mL x 10 vials',
            'unit': 'Set',
            'par': '2',
            'min': '1',
            'max': '4',
            'storage_temp': '-20°C',
            'storage_loc': 'Coag Freezer',
            'analyzer': 'STAGO Compact',
            'test': 'PT/INR QC',
            'critical': 'YES',
            'notes': 'Check green binder'
        },
        {
            'item_num': 'KT005',
            'desc': 'COVID-19 Rapid Test Kit',
            'manufacturer': 'Abbott',
            'mfr_cat': 'BINAXNOW',
            'material': '',
            'supplier': '',
            'package_size': '40 tests/box',
            'unit': 'Box',
            'par': '5',
            'min': '2',
            'max': '10',
            'storage_temp': 'Room Temp',
            'storage_loc': 'POC Storage',
            'analyzer': 'N/A',
            'test': 'COVID-19 Antigen',
            'critical': 'YES',
            'notes': 'Check expiration dates'
        }
    ]
    
    # MISCELLANEOUS - Detailed items
    misc_items = [
        {
            'item_num': 'MS001',
            'desc': 'Digital Thermometer - Refrigerator/Freezer',
            'manufacturer': 'VWR',
            'mfr_cat': '35519-040',
            'material': '',
            'supplier': '10311248',
            'package_size': '1 unit',
            'unit': 'Each',
            'par': '2',
            'min': '1',
            'max': '4',
            'storage_temp': 'N/A',
            'storage_loc': 'Various',
            'analyzer': 'N/A',
            'test': 'Temperature Monitor',
            'critical': 'YES',
            'notes': 'AAA batteries, check daily'
        },
        {
            'item_num': 'MS002',
            'desc': 'Thermal Paper - MEDTOX Scanner',
            'manufacturer': 'MEDTOX',
            'mfr_cat': 'TP-110',
            'material': '',
            'supplier': '10333259',
            'package_size': '10 rolls/box',
            'unit': 'Box',
            'par': '2',
            'min': '1',
            'max': '4',
            'storage_temp': 'Room Temp',
            'storage_loc': 'Supply Cabinet',
            'analyzer': 'MEDTOX Scanner',
            'test': 'N/A',
            'critical': 'NO'
        },
        {
            'item_num': 'MS003',
            'desc': 'Variable Volume Pipette 20-200µL',
            'manufacturer': 'Eppendorf',
            'mfr_cat': '3123000055',
            'material': '',
            'supplier': '10333263',
            'package_size': '1 unit',
            'unit': 'Each',
            'par': '2',
            'min': '1',
            'max': '3',
            'storage_temp': 'Room Temp',
            'storage_loc': 'Pipette Rack',
            'analyzer': 'N/A',
            'test': 'Manual Testing',
            'critical': 'YES',
            'notes': 'Calibrate annually'
        },
        {
            'item_num': 'MS004',
            'desc': 'Hematek Slide Stainer Cleaning Solution',
            'manufacturer': 'Siemens',
            'mfr_cat': 'T402',
            'material': '',
            'supplier': '',
            'package_size': '1L bottle',
            'unit': 'Each',
            'par': '2',
            'min': '1',
            'max': '4',
            'storage_temp': 'Room Temp',
            'storage_loc': 'Hematology',
            'analyzer': 'Hematek Stainer',
            'test': 'Maintenance',
            'critical': 'NO',
            'notes': 'Daily QC slide, record in maintenance log'
        },
        {
            'item_num': 'MS005',
            'desc': 'Nitrile Exam Gloves - Small',
            'manufacturer': 'Kimberly-Clark',
            'mfr_cat': '55081',
            'material': '',
            'supplier': '',
            'package_size': '100/box',
            'unit': 'Box',
            'par': '20',
            'min': '10',
            'max': '40',
            'storage_temp': 'Room Temp',
            'storage_loc': 'PPE Station',
            'analyzer': 'N/A',
            'test': 'N/A',
            'critical': 'YES',
            'notes': 'Purple nitrile'
        },
        {
            'item_num': 'MS006',
            'desc': 'Nitrile Exam Gloves - Medium',
            'manufacturer': 'Kimberly-Clark',
            'mfr_cat': '55082',
            'material': '',
            'supplier': '',
            'package_size': '100/box',
            'unit': 'Box',
            'par': '30',
            'min': '15',
            'max': '60',
            'storage_temp': 'Room Temp',
            'storage_loc': 'PPE Station',
            'analyzer': 'N/A',
            'test': 'N/A',
            'critical': 'YES',
            'notes': 'Purple nitrile'
        },
        {
            'item_num': 'MS007',
            'desc': 'Nitrile Exam Gloves - Large',
            'manufacturer': 'Kimberly-Clark',
            'mfr_cat': '55083',
            'material': '',
            'supplier': '',
            'package_size': '100/box',
            'unit': 'Box',
            'par': '25',
            'min': '12',
            'max': '50',
            'storage_temp': 'Room Temp',
            'storage_loc': 'PPE Station',
            'analyzer': 'N/A',
            'test': 'N/A',
            'critical': 'YES',
            'notes': 'Purple nitrile'
        },
        {
            'item_num': 'MS008',
            'desc': 'Sharps Container - 5 Quart',
            'manufacturer': 'BD',
            'mfr_cat': '305557',
            'material': '',
            'supplier': '',
            'package_size': '20/case',
            'unit': 'Case',
            'par': '2',
            'min': '1',
            'max': '4',
            'storage_temp': 'Room Temp',
            'storage_loc': 'Supply Room',
            'analyzer': 'N/A',
            'test': 'N/A',
            'critical': 'YES',
            'notes': 'Red containers'
        },
        {
            'item_num': 'MS009',
            'desc': 'QC Log Book - Chemistry',
            'manufacturer': 'Lab Armor',
            'mfr_cat': 'QC-CHEM',
            'material': '',
            'supplier': '',
            'package_size': '1 book',
            'unit': 'Each',
            'par': '2',
            'min': '1',
            'max': '3',
            'storage_temp': 'N/A',
            'storage_loc': 'QC Area',
            'analyzer': 'N/A',
            'test': 'Documentation',
            'critical': 'YES',
            'notes': 'Green binder'
        },
        {
            'item_num': 'MS010',
            'desc': 'Maintenance Log Book - Hematology',
            'manufacturer': 'Lab Armor',
            'mfr_cat': 'MX-HEM',
            'material': '',
            'supplier': '',
            'package_size': '1 book',
            'unit': 'Each',
            'par': '2',
            'min': '1',
            'max': '3',
            'storage_temp': 'N/A',
            'storage_loc': 'Hematology',
            'analyzer': 'N/A',
            'test': 'Documentation',
            'critical': 'YES',
            'notes': 'Green binder'
        }
    ]
    
    # Create category sheets with detailed items
    categories = {
        'CHEMISTRY': chemistry_items,
        'HEMATOLOGY': hematology_items,
        'URINALYSIS': urinalysis_items,
        'KITS': kits_items,
        'MISCELLANEOUS': misc_items
    }
    
    for category, items in categories.items():
        sheet = wb.create_sheet(title=category)
        
        # Add headers
        for col_idx, col_info in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col_idx, value=col_info['name'])
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )
            sheet.column_dimensions[get_column_letter(col_idx)].width = col_info['width']
        
        # Add items with all details
        for row_idx, item in enumerate(items, 2):
            # Map all fields
            sheet.cell(row=row_idx, column=1, value=item.get('item_num', ''))
            sheet.cell(row=row_idx, column=2, value=item.get('desc', ''))
            sheet.cell(row=row_idx, column=3, value=item.get('manufacturer', ''))
            sheet.cell(row=row_idx, column=4, value=item.get('mfr_cat', ''))
            sheet.cell(row=row_idx, column=5, value=item.get('material', ''))
            sheet.cell(row=row_idx, column=6, value=item.get('supplier', ''))
            sheet.cell(row=row_idx, column=7, value=item.get('onelink', ''))
            sheet.cell(row=row_idx, column=8, value=item.get('package_size', ''))
            sheet.cell(row=row_idx, column=9, value=item.get('unit', ''))
            sheet.cell(row=row_idx, column=10, value=item.get('par', ''))
            sheet.cell(row=row_idx, column=11, value=item.get('min', ''))
            sheet.cell(row=row_idx, column=12, value=item.get('max', ''))
            sheet.cell(row=row_idx, column=13, value=item.get('hand_count', ''))
            sheet.cell(row=row_idx, column=14, value=item.get('req_qty', ''))
            sheet.cell(row=row_idx, column=15, value=item.get('reorder_point', ''))
            sheet.cell(row=row_idx, column=20, value=item.get('storage_temp', ''))
            sheet.cell(row=row_idx, column=21, value=item.get('storage_loc', ''))
            sheet.cell(row=row_idx, column=22, value=item.get('analyzer', ''))
            sheet.cell(row=row_idx, column=23, value=item.get('test', ''))
            sheet.cell(row=row_idx, column=24, value=item.get('critical', ''))
            sheet.cell(row=row_idx, column=27, value=datetime.now().strftime('%Y-%m-%d'))
            sheet.cell(row=row_idx, column=28, value='Initial Setup')
            sheet.cell(row=row_idx, column=29, value=item.get('notes', ''))
            
            # Calculate status based on hand count
            if item.get('hand_count'):
                try:
                    count = int(item.get('hand_count', 0))
                    min_stock = int(item.get('min', 0)) if item.get('min') else 0
                    
                    if count == 0:
                        status = 'OUT OF STOCK'
                        color = 'FF0000'  # Red
                    elif count <= min_stock:
                        status = 'CRITICAL LOW'
                        color = 'FF6600'  # Orange
                    elif count < 10:
                        status = 'LOW STOCK'
                        color = 'FFCC00'  # Yellow
                    else:
                        status = 'OK'
                        color = '00FF00'  # Green
                    
                    sheet.cell(row=row_idx, column=16, value=status)
                    sheet.cell(row=row_idx, column=16).fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )
                except:
                    pass
            
            # Highlight critical items
            if item.get('critical') == 'YES':
                sheet.cell(row=row_idx, column=24).font = Font(bold=True, color="FF0000")
            
            # Special highlighting for ALT reagents
            if 'ALT' in item.get('desc', '') and 'EXPIRE' in item.get('notes', ''):
                for col in range(1, 31):
                    sheet.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFE5E5", end_color="FFE5E5", fill_type="solid"
                    )
                sheet.cell(row=row_idx, column=30, value='URGENT - REDISTRIBUTE!')
                sheet.cell(row=row_idx, column=30).font = Font(bold=True, color="FF0000")
            
            # Add borders
            for col in range(1, 31):
                sheet.cell(row=row_idx, column=col).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Freeze top row and first two columns
        sheet.freeze_panes = 'C2'
        
        # Add data validation
        # Location dropdown
        location_dv = DataValidation(
            type="list",
            formula1='"MOB,AUC,BOTH,MOB→AUC,AUC→MOB"',
            allow_blank=True
        )
        location_dv.add(f'S2:S1000')
        sheet.add_data_validation(location_dv)
        
        # Critical item dropdown
        critical_dv = DataValidation(
            type="list",
            formula1='"YES,NO"',
            allow_blank=True
        )
        critical_dv.add(f'X2:X1000')
        sheet.add_data_validation(critical_dv)
        
        # Status dropdown
        status_dv = DataValidation(
            type="list",
            formula1='"OK,LOW STOCK,CRITICAL LOW,OUT OF STOCK,EXPIRED,ON ORDER"',
            allow_blank=True
        )
        status_dv.add(f'P2:P1000')
        sheet.add_data_validation(status_dv)
    
    # Create enhanced summary dashboard
    summary = wb.create_sheet(title='DASHBOARD', index=0)
    
    # Title section
    summary.merge_cells('A1:J1')
    title_cell = summary['A1']
    title_cell.value = "LARGO LAB DETAILED INVENTORY MANAGEMENT SYSTEM"
    title_cell.font = Font(bold=True, size=20, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[1].height = 40
    
    # Timestamp
    summary['A2'] = "Last Updated:"
    summary['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    summary['B2'].font = Font(bold=True)
    summary['C2'] = "Updated By:"
    summary['D2'] = "System Administrator"
    
    # Critical alerts
    row = 4
    summary.merge_cells(f'A{row}:J{row}')
    summary[f'A{row}'] = "🚨 CRITICAL ALERTS - IMMEDIATE ACTION REQUIRED"
    summary[f'A{row}'].font = Font(bold=True, size=16, color="FF0000")
    summary[f'A{row}'].fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
    
    row += 2
    alerts = [
        "🔴 ALT REAGENT PACKS (25 TOTAL) EXPIRE OCTOBER 31 - REDISTRIBUTE TO OTHER KAISER LOCATIONS IMMEDIATELY",
        "🟡 MULTIPLE SUPPLIER IDs NEED VERIFICATION - MAXWELL BOOKER IDENTIFIED ERRORS",
        "🟡 MEDTOX QC MUST BE LOGGED IN CERNER - CHECK MAINTENANCE LOG",
        "🟢 PNEUMATIC TUBE SYSTEM NOW FUNCTIONAL - UPDATE PROCEDURES",
        "🟢 CBC's CAN BE RUN AT MOB LAB - UPDATE WORKFLOW",
        "🔵 URINE PROCESSING: LORRAINE (PRIMARY), MIMI (BACKUP WHEN LORRAINE OUT)"
    ]
    
    for alert in alerts:
        summary[f'A{row}'] = alert
        summary.merge_cells(f'A{row}:J{row}')
        if alert.startswith('🔴'):
            summary[f'A{row}'].font = Font(bold=True, color="FF0000", size=12)
        elif alert.startswith('🟡'):
            summary[f'A{row}'].font = Font(bold=True, color="FF6600", size=12)
        else:
            summary[f'A{row}'].font = Font(size=12)
        row += 1
    
    # Lab information
    row += 2
    summary.merge_cells(f'A{row}:J{row}')
    summary[f'A{row}'] = "LAB LOCATIONS & CAPABILITIES"
    summary[f'A{row}'].font = Font(bold=True, size=14)
    
    row += 2
    lab_info = [
        ['Location', 'Tests Available', 'Key Equipment', 'Staff Lead'],
        ['MOB Lab (Core)', 'CBC, Chemistry, UA', 'Roche c303, Sysmex XN-1000', 'Lorraine'],
        ['AUC Lab (STAT)', 'CBC, Chemistry, UA, Coag', 'Roche c503, STAGO Compact', 'Ingrid']
    ]
    
    for r_idx, row_data in enumerate(lab_info):
        for c_idx, value in enumerate(row_data):
            cell = summary.cell(row=row + r_idx, column=c_idx + 1, value=value)
            if r_idx == 0:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Inventory statistics
    row += 6
    summary.merge_cells(f'A{row}:J{row}')
    summary[f'A{row}'] = "INVENTORY STATISTICS BY CATEGORY"
    summary[f'A{row}'].font = Font(bold=True, size=14)
    
    row += 2
    stats_headers = ['Category', 'Total Items', 'Critical Items', 'Out of Stock', 'Low Stock', 'Expiring Soon', 'Items OK', 'Fill Rate %']
    for c_idx, header in enumerate(stats_headers):
        cell = summary.cell(row=row, column=c_idx + 1, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Add formulas for each category
    for idx, category in enumerate(['CHEMISTRY', 'HEMATOLOGY', 'URINALYSIS', 'KITS', 'MISCELLANEOUS'], row + 1):
        summary.cell(row=idx, column=1, value=category)
        summary.cell(row=idx, column=2, value=f'=COUNTA({category}!A:A)-1')
        summary.cell(row=idx, column=3, value=f'=COUNTIF({category}!X:X,"YES")')
        summary.cell(row=idx, column=4, value=f'=COUNTIF({category}!P:P,"OUT OF STOCK")')
        summary.cell(row=idx, column=5, value=f'=COUNTIF({category}!P:P,"*LOW*")')
        summary.cell(row=idx, column=6, value=f'=COUNTIF({category}!AD:AD,"*EXPIRE*")')
        summary.cell(row=idx, column=7, value=f'=COUNTIF({category}!P:P,"OK")')
        summary.cell(row=idx, column=8, value=f'=G{idx}/B{idx}*100')
        summary.cell(row=idx, column=8).number_format = '0.0"%"'
        
        for col in range(1, 9):
            summary.cell(row=idx, column=col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Staff quick reference
    row += 8
    summary.merge_cells(f'A{row}:J{row}')
    summary[f'A{row}'] = "QUICK REFERENCE - STAFF RESPONSIBILITIES"
    summary[f'A{row}'].font = Font(bold=True, size=14)
    
    row += 2
    staff_info = [
        "• Inventory Manager (Primary): LORRAINE - Overall inventory, urine processing",
        "• Inventory Manager (Secondary): INGRID - Backup, coordinates overflow between labs",
        "• Urine Processing Backup: MIMI - When Lorraine is out",
        "• Supply Ordering: NATHANIEL BURMEISTER - Max quantities, new lots",
        "• Supplier ID Verification: MAXWELL BOOKER - Found errors, helping verify",
        "• QC Documentation: ALL TECHS - Must log in Cerner AND green binders",
        "• Overflow Coordination: Send to AUC when MOB overwhelmed, use Teams chat"
    ]
    
    for info in staff_info:
        summary[f'A{row}'] = info
        summary.merge_cells(f'A{row}:J{row}')
        summary[f'A{row}'].font = Font(size=11)
        row += 1
    
    # Instructions
    row += 2
    summary.merge_cells(f'A{row}:J{row}')
    summary[f'A{row}'] = "DAILY PROCEDURES"
    summary[f'A{row}'].font = Font(bold=True, size=14)
    
    row += 2
    procedures = [
        "1. START OF SHIFT:",
        "   • Check DASHBOARD for critical alerts",
        "   • Review items marked OUT OF STOCK or LOW STOCK",
        "   • Run QC for all analyzers, log in Cerner AND binders",
        "",
        "2. DURING SHIFT:",
        "   • Update HAND COUNT when using supplies",
        "   • Note any supplier ID discrepancies in NOTES column",
        "   • Monitor refrigerator/freezer temperatures",
        "",
        "3. RECEIVING SUPPLIES:",
        "   • Verify SUPPLIER ID matches packing slip",
        "   • Update HAND COUNT immediately",
        "   • Enter EXPIRATION DATE and LOT NUMBER",
        "   • Change STATUS to 'OK' if previously out",
        "",
        "4. END OF SHIFT:",
        "   • Save spreadsheet",
        "   • Communicate any critical issues to next shift",
        "   • Update Teams channel with any concerns"
    ]
    
    for proc in procedures:
        summary[f'A{row}'] = proc
        if proc and not proc.startswith('   '):
            summary[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # PTO reminder
    row += 2
    summary.merge_cells(f'A{row}:J{row}')
    summary[f'A{row}'] = "⚠️ REMINDER: Q4 PTO requests due TODAY - Seniority rules apply per labor contract"
    summary[f'A{row}'].font = Font(bold=True, color="FF0000", size=12)
    summary[f'A{row}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # Adjust column widths for dashboard
    for col in range(1, 11):
        summary.column_dimensions[get_column_letter(col)].width = 15
    
    # Create EXPIRING_ITEMS sheet
    expiring = wb.create_sheet(title='EXPIRING_ITEMS')
    
    # Headers
    exp_headers = [
        'PRIORITY', 'ITEM #', 'DESCRIPTION', 'QUANTITY', 'LOCATION', 
        'EXPIRATION DATE', 'DAYS UNTIL EXPIRY', 'LOT NUMBER', 
        'ACTION PLAN', 'CONTACT', 'STATUS'
    ]
    
    for col_idx, header in enumerate(exp_headers, 1):
        cell = expiring.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Add ALT reagent as priority item
    expiring['A2'] = '🔴 URGENT'
    expiring['B2'] = 'CH001'
    expiring['C2'] = 'ALT Reagent Packs'
    expiring['D2'] = '25'
    expiring['E2'] = 'MOB/AUC'
    expiring['F2'] = '2025-10-31'
    expiring['G2'] = '=F2-TODAY()'
    expiring['H2'] = 'Multiple'
    expiring['I2'] = 'Contact other Kaiser locations for redistribution'
    expiring['J2'] = 'John F Ekpe, Ingrid Z Benitez-Ruiz'
    expiring['K2'] = 'PENDING'
    
    # Highlight entire row
    for col in range(1, 12):
        expiring.cell(row=2, column=col).fill = PatternFill(
            start_color="FFE5E5", end_color="FFE5E5", fill_type="solid"
        )
        expiring.cell(row=2, column=col).border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Adjust column widths
    exp_widths = {
        'A': 12, 'B': 10, 'C': 40, 'D': 12, 'E': 15,
        'F': 20, 'G': 20, 'H': 15, 'I': 50, 'J': 30, 'K': 15
    }
    for col, width in exp_widths.items():
        expiring.column_dimensions[col].width = width
    
    expiring.freeze_panes = 'A2'
    
    return wb

def main():
    print("Creating DETAILED Lab Inventory Management System")
    print("=" * 70)
    
    # Create the detailed inventory
    wb = create_detailed_inventory()
    
    # Save the file
    downloads_path = Path("/Users/ugochi141/Downloads")
    filename = "LARGO_LAB_DETAILED_INVENTORY_SYSTEM.xlsx"
    filepath = downloads_path / filename
    wb.save(filepath)
    
    # Also save to project
    project_path = Path("/Users/ugochi141/Desktop/LabAutomation/data/inventory")
    project_file = project_path / filename
    wb.save(project_file)
    
    print(f"✅ Detailed inventory system created: {filename}")
    print(f"   Downloads: {filepath}")
    print(f"   Project: {project_file}")
    
    print("\n" + "=" * 70)
    print("📊 DETAILED INVENTORY FEATURES:")
    print("=" * 70)
    print("✓ 30 columns of detailed tracking information")
    print("✓ Item numbers for easy reference")
    print("✓ Manufacturer and catalog numbers")
    print("✓ Package sizes and units of measure")
    print("✓ PAR levels, min/max stock levels")
    print("✓ Storage temperatures and locations")
    print("✓ Analyzer/equipment assignments")
    print("✓ Test/procedure associations")
    print("✓ Critical item flagging")
    print("✓ Complete staff assignments")
    print("✓ Detailed dashboard with statistics")
    print("✓ Expiring items tracking with action plans")
    print("✓ Data validation dropdowns")
    print("✓ Color-coded status indicators")
    print("✓ ALT reagent crisis highlighted throughout")
    
    print("\n🎯 IMMEDIATE ACTIONS:")
    print("1. Open the file and review DASHBOARD tab")
    print("2. Check EXPIRING_ITEMS for ALT reagents")
    print("3. Fill in missing HAND COUNTS")
    print("4. Verify SUPPLIER IDs")
    print("5. Upload to Teams/SharePoint")
    print("6. Train staff on all 30 fields")

if __name__ == "__main__":
    main()



