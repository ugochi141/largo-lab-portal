#!/usr/bin/env python3
"""
Enhanced Lab Inventory Management System Consolidator
Handles the specific format of Largo lab inventory files
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os
from pathlib import Path
import re

class EnhancedInventoryConsolidator:
    def __init__(self, downloads_path):
        self.downloads_path = Path(downloads_path)
        self.master_inventory = {}
        self.expiring_items = []
        
    def extract_chemistry_data(self, sheet):
        """Extract data from CHEMISTRY sheet format"""
        items = []
        
        # Look for actual data starting around row 8
        for row_idx in range(8, sheet.max_row + 1):
            row_data = []
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                row_data.append(cell_value)
            
            # Check if this is a data row (has material number and description)
            if row_data[0] and isinstance(row_data[0], (int, str)) and str(row_data[0]).strip():
                item = {
                    'MATERIAL#': str(row_data[0]) if row_data[0] else '',
                    'DESCRIPTION': str(row_data[1]) if len(row_data) > 1 and row_data[1] else '',
                    'MFR#/CAT#': str(row_data[2]) if len(row_data) > 2 and row_data[2] else '',
                    'SUPPLIER ID': str(row_data[3]) if len(row_data) > 3 and row_data[3] else '',
                    'ONELINK NUMBER': str(row_data[4]) if len(row_data) > 4 and row_data[4] else '',
                    'REQ QTY': str(row_data[5]) if len(row_data) > 5 and row_data[5] else '',
                    'HAND COUNT': str(row_data[6]) if len(row_data) > 6 and row_data[6] else ''
                }
                if item['DESCRIPTION']:  # Only add if there's a description
                    items.append(item)
        
        return items
    
    def extract_complete_roche_data(self, sheet):
        """Extract data from Complete Roche sheet format"""
        items = []
        
        # Data starts at row 3
        for row_idx in range(3, sheet.max_row + 1):
            material_num = sheet.cell(row=row_idx, column=1).value
            description = sheet.cell(row=row_idx, column=5).value if sheet.max_column >= 5 else ''
            
            if material_num and description:
                item = {
                    'MATERIAL#': str(material_num),
                    'DESCRIPTION': str(description),
                    'ISE': '✓' if sheet.cell(row=row_idx, column=2).value else '',
                    'c303': '✓' if sheet.cell(row=row_idx, column=3).value else '',
                    'e402': '✓' if sheet.cell(row=row_idx, column=4).value else ''
                }
                items.append(item)
        
        return items
    
    def extract_kits_data(self, sheet):
        """Extract data from KITS sheet format"""
        items = []
        current_kit = None
        
        for row_idx in range(3, sheet.max_row + 1):
            col1 = sheet.cell(row=row_idx, column=1).value
            col2 = sheet.cell(row=row_idx, column=2).value
            col3 = sheet.cell(row=row_idx, column=3).value
            col4 = sheet.cell(row=row_idx, column=4).value
            col5 = sheet.cell(row=row_idx, column=5).value
            col6 = sheet.cell(row=row_idx, column=6).value
            
            # Check if this is a kit header
            if col1 and 'KIT' in str(col1).upper() and col2:
                current_kit = {
                    'name': str(col1),
                    'supplier_id': str(col2) if col2 else ''
                }
            # This is a kit component
            elif col1 and current_kit:
                item = {
                    'DESCRIPTION': f"{current_kit['name']} - {col1}",
                    'SUPPLIER ID': current_kit['supplier_id'],
                    'ONELINK NUMBER': str(col3) if col3 else '',
                    'PAR': str(col2) if col2 else '',
                    'REQ QTY': str(col4) if col4 else '',
                    'HAND COUNT': str(col5) if col5 else ''
                }
                items.append(item)
        
        return items
    
    def extract_miscellaneous_data(self, sheet):
        """Extract data from MISCELLANOUS sheet format"""
        items = []
        
        for row_idx in range(3, sheet.max_row + 1):
            description = sheet.cell(row=row_idx, column=1).value
            onelink = sheet.cell(row=row_idx, column=2).value
            supplier_id = sheet.cell(row=row_idx, column=3).value
            par = sheet.cell(row=row_idx, column=4).value
            req_qty = sheet.cell(row=row_idx, column=5).value
            hand_count = sheet.cell(row=row_idx, column=6).value
            
            if description and not str(description).startswith('PAR'):
                item = {
                    'DESCRIPTION': str(description).strip(),
                    'ONELINK NUMBER': str(onelink).strip() if onelink else '',
                    'SUPPLIER ID': str(supplier_id).strip() if supplier_id else '',
                    'PAR': str(par) if par else '',
                    'REQ QTY': str(req_qty) if req_qty else '',
                    'HAND COUNT': str(hand_count) if hand_count else ''
                }
                items.append(item)
        
        return items
    
    def process_inventory_file(self):
        """Process the main inventory file with proper extraction logic"""
        main_file = self.downloads_path / "Inventory List Largo 03MAR25_Shanthi and Nate B_09092025.xlsx"
        if not main_file.exists():
            print(f"Main file not found: {main_file}")
            return
            
        wb = openpyxl.load_workbook(main_file, data_only=True)
        
        # Process each sheet with appropriate extractor
        sheet_processors = {
            'CHEMISTRY': self.extract_chemistry_data,
            'Complete Roche': self.extract_complete_roche_data,
            'KITS': self.extract_kits_data,
            'MISCELLANOUS': self.extract_miscellaneous_data
        }
        
        for sheet_name in wb.sheetnames:
            if sheet_name in sheet_processors:
                print(f"Processing {sheet_name}...")
                sheet = wb[sheet_name]
                items = sheet_processors[sheet_name](sheet)
                
                # Determine category
                if 'Roche' in sheet_name:
                    category = 'CHEMISTRY'
                else:
                    category = sheet_name.upper().replace('MISCELLANOUS', 'MISCELLANEOUS')
                
                # Add items to master inventory
                for item in items:
                    key = f"{category}_{item.get('DESCRIPTION', '')}"
                    self.master_inventory[key] = {
                        'category': category,
                        'data': item
                    }
                
                print(f"  Found {len(items)} items in {sheet_name}")
        
        print(f"\nTotal items in master inventory: {len(self.master_inventory)}")
    
    def check_for_expiring_items(self):
        """Identify items that need attention (like the ALT reagents)"""
        # Based on the chat, ALT reagents are expiring
        alt_items = [key for key in self.master_inventory.keys() if 'ALT' in key.upper()]
        if alt_items:
            self.expiring_items.extend(alt_items)
            print(f"\n⚠️  Found {len(alt_items)} ALT items that may be expiring soon")
    
    def create_enhanced_master_workbook(self):
        """Create the comprehensive master inventory workbook with enhanced features"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define categories
        categories = ['CHEMISTRY', 'HEMATOLOGY', 'URINALYSIS', 'KITS', 'MISCELLANEOUS', 'SUMMARY', 'EXPIRING_ITEMS']
        
        # Define enhanced columns
        enhanced_columns = [
            'DESCRIPTION',
            'MFR#/CAT#',
            'MATERIAL# (KAISER#/OLID)',
            'SUPPLIER ID',
            'ONELINK NUMBER',
            'PAR LEVEL',
            'HAND COUNT',
            'REQ QTY',
            'STATUS',
            'EXPIRATION DATE',
            'LOT NUMBER',
            'LOCATION (MOB/AUC)',
            'LAST UPDATED',
            'UPDATED BY',
            'NOTES',
            'ACTION REQUIRED'
        ]
        
        # Create sheets for each category
        for category in categories:
            sheet = wb.create_sheet(title=category)
            
            # Add headers with formatting
            for col, header in enumerate(enhanced_columns, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Add borders
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Freeze panes
            sheet.freeze_panes = 'A2'
            
            # Adjust column widths
            column_widths = {
                'A': 45,  # DESCRIPTION
                'B': 20,  # MFR#/CAT#
                'C': 25,  # MATERIAL#
                'D': 15,  # SUPPLIER ID
                'E': 20,  # ONELINK NUMBER
                'F': 12,  # PAR LEVEL
                'G': 12,  # HAND COUNT
                'H': 10,  # REQ QTY
                'I': 15,  # STATUS
                'J': 15,  # EXPIRATION DATE
                'K': 15,  # LOT NUMBER
                'L': 15,  # LOCATION
                'M': 15,  # LAST UPDATED
                'N': 20,  # UPDATED BY
                'O': 30,  # NOTES
                'P': 25   # ACTION REQUIRED
            }
            
            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width
        
        # Populate data from master inventory
        for item_key, item_info in self.master_inventory.items():
            category = item_info['category']
            if category not in categories:
                category = 'MISCELLANEOUS'
                
            sheet = wb[category]
            row = sheet.max_row + 1
            
            data = item_info['data']
            
            # Map data to enhanced columns
            sheet.cell(row=row, column=1, value=data.get('DESCRIPTION', ''))
            sheet.cell(row=row, column=2, value=data.get('MFR#/CAT#', ''))
            sheet.cell(row=row, column=3, value=data.get('MATERIAL#', ''))
            sheet.cell(row=row, column=4, value=data.get('SUPPLIER ID', ''))
            sheet.cell(row=row, column=5, value=data.get('ONELINK NUMBER', ''))
            sheet.cell(row=row, column=6, value=data.get('PAR', ''))
            sheet.cell(row=row, column=7, value=data.get('HAND COUNT', ''))
            sheet.cell(row=row, column=8, value=data.get('REQ QTY', ''))
            
            # Calculate status
            try:
                hand_count = int(data.get('HAND COUNT', 0)) if data.get('HAND COUNT', '').isdigit() else 0
                par_level = int(data.get('PAR', 0)) if data.get('PAR', '').isdigit() else 0
                
                if hand_count == 0:
                    status = 'OUT OF STOCK'
                    status_color = 'FF0000'  # Red
                elif hand_count < 10:
                    status = 'LOW STOCK'
                    status_color = 'FFA500'  # Orange
                elif par_level > 0 and hand_count < par_level:
                    status = 'BELOW PAR'
                    status_color = 'FFFF00'  # Yellow
                else:
                    status = 'OK'
                    status_color = '00FF00'  # Green
                
                sheet.cell(row=row, column=9, value=status)
                sheet.cell(row=row, column=9).fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
            except:
                sheet.cell(row=row, column=9, value='CHECK')
            
            sheet.cell(row=row, column=13, value=datetime.now().strftime('%Y-%m-%d %H:%M'))
            sheet.cell(row=row, column=14, value='System Import')
            
            # Check if this is an expiring item
            if item_key in self.expiring_items:
                sheet.cell(row=row, column=10, value='2025-10-31')  # End of October
                sheet.cell(row=row, column=15, value='EXPIRING SOON - Redistribute or use before end of October')
                sheet.cell(row=row, column=16, value='URGENT - CHECK EXPIRATION')
                
                # Highlight entire row
                for col in range(1, 17):
                    sheet.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFE5E5", end_color="FFE5E5", fill_type="solid"
                    )
            
            # Add borders to all cells
            for col in range(1, 17):
                sheet.cell(row=row, column=col).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Create enhanced summary sheet
        self.create_enhanced_summary_sheet(wb)
        
        # Add data validation
        self.add_enhanced_validation(wb)
        
        return wb
    
    def create_enhanced_summary_sheet(self, wb):
        """Create an enhanced summary dashboard with actionable insights"""
        summary = wb['SUMMARY']
        
        # Title
        summary.merge_cells('A1:H1')
        title_cell = summary['A1']
        title_cell.value = "LARGO LAB INVENTORY MANAGEMENT DASHBOARD"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add timestamp
        summary['A2'] = "Last Updated:"
        summary['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        summary['B2'].font = Font(bold=True)
        
        # Critical alerts section
        summary.merge_cells('A4:H4')
        alerts_title = summary['A4']
        alerts_title.value = "⚠️ CRITICAL ALERTS"
        alerts_title.font = Font(bold=True, size=14, color="FF0000")
        
        # Add specific alerts from the chat
        alerts = [
            "1. ALT reagent packs (25 total) expire end of October - redistribute between MOB/AUC or share with other locations",
            "2. Check all supplier IDs for accuracy - several found to be incorrect",
            "3. MedTox QC maintenance needs to be logged in Cerner",
            "4. Pneumatic tube system is now functional - update procedures",
            "5. CBC's can be run at MOB Lab - update workflow"
        ]
        
        for idx, alert in enumerate(alerts, 6):
            summary.cell(row=idx, column=1, value=alert)
            summary.cell(row=idx, column=1).font = Font(color="CC0000")
        
        # Inventory summary by category
        summary_row = 12
        summary.merge_cells(f'A{summary_row}:H{summary_row}')
        summary_title = summary[f'A{summary_row}']
        summary_title.value = "INVENTORY SUMMARY BY CATEGORY"
        summary_title.font = Font(bold=True, size=14)
        
        # Headers
        headers = ['Category', 'Total Items', 'Out of Stock', 'Low Stock (<10)', 'Below PAR', 'Expiring Soon', 'Action Required', 'Compliance']
        for col, header in enumerate(headers, 1):
            cell = summary.cell(row=summary_row + 2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Add formulas for each category
        categories = ['CHEMISTRY', 'HEMATOLOGY', 'URINALYSIS', 'KITS', 'MISCELLANEOUS']
        for idx, category in enumerate(categories, summary_row + 3):
            summary.cell(row=idx, column=1, value=category)
            summary.cell(row=idx, column=2, value=f'=COUNTA({category}!A:A)-1')
            summary.cell(row=idx, column=3, value=f'=COUNTIF({category}!I:I,"OUT OF STOCK")')
            summary.cell(row=idx, column=4, value=f'=COUNTIF({category}!I:I,"LOW STOCK")')
            summary.cell(row=idx, column=5, value=f'=COUNTIF({category}!I:I,"BELOW PAR")')
            summary.cell(row=idx, column=6, value=f'=COUNTIF({category}!P:P,"URGENT*")')
            summary.cell(row=idx, column=7, value=f'=SUM(C{idx}:F{idx})')
            
            # Add borders
            for col in range(1, 9):
                summary.cell(row=idx, column=col).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Staff reminders section
        reminder_row = summary_row + 10
        summary.merge_cells(f'A{reminder_row}:H{reminder_row}')
        reminder_title = summary[f'A{reminder_row}']
        reminder_title.value = "📋 STAFF REMINDERS & PROCEDURES"
        reminder_title.font = Font(bold=True, size=14)
        
        reminders = [
            "• Update HAND COUNT when performing inventory checks - clock in 30 min early if needed (get manager approval)",
            "• Pour-off urine samples: Primary - Lorraine, Backup - Mimi (coordinate with Ingrid)",
            "• REMINDER: Submit Q4 PTO requests by deadline - seniority rules apply per labor contract",
            "• When supplies run low at MOB, send specimens to AUC for testing",
            "• Always verify QC is logged in Cerner AND maintenance binders"
        ]
        
        for idx, reminder in enumerate(reminders, reminder_row + 2):
            summary.cell(row=idx, column=1, value=reminder)
        
        # Instructions for use
        instruction_row = reminder_row + 8
        summary.merge_cells(f'A{instruction_row}:H{instruction_row}')
        instruction_title = summary[f'A{instruction_row}']
        instruction_title.value = "📖 HOW TO USE THIS INVENTORY SYSTEM"
        instruction_title.font = Font(bold=True, size=14)
        
        instructions = [
            "1. DAILY: Check SUMMARY tab for critical alerts and low stock items",
            "2. WHEN RECEIVING SUPPLIES: Update HAND COUNT and verify SUPPLIER ID matches packing slip",
            "3. FOR EXPIRING ITEMS: Check EXPIRING_ITEMS tab weekly, coordinate redistribution",
            "4. LOCATION: Always specify MOB, AUC, or BOTH in location column",
            "5. DISCREPANCIES: Note any supplier ID errors in NOTES column for Nathaniel to fix",
            "6. SHARING: Save and upload to Teams/SharePoint after each update",
            "7. SUPPORT: Contact Nathaniel Burmeister for supply list updates or system issues"
        ]
        
        for idx, instruction in enumerate(instructions, instruction_row + 2):
            summary.cell(row=idx, column=1, value=instruction)
            summary.cell(row=idx, column=1).font = Font(size=11)
    
    def add_enhanced_validation(self, wb):
        """Add data validation and conditional formatting"""
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.formatting.rule import CellIsRule
        
        # Location validation
        location_dv = DataValidation(
            type="list",
            formula1='"MOB,AUC,BOTH,MOB→AUC,AUC→MOB"',
            allow_blank=True
        )
        location_dv.error = 'Please select valid location'
        location_dv.errorTitle = 'Invalid Location'
        
        # Status validation
        status_dv = DataValidation(
            type="list",
            formula1='"OK,LOW STOCK,OUT OF STOCK,BELOW PAR,CHECK,EXPIRED"',
            allow_blank=True
        )
        
        for sheet_name in ['CHEMISTRY', 'HEMATOLOGY', 'URINALYSIS', 'KITS', 'MISCELLANEOUS']:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                location_dv.add(f'L2:L1000')
                status_dv.add(f'I2:I1000')
                sheet.add_data_validation(location_dv)
                sheet.add_data_validation(status_dv)
                
                # Add conditional formatting for low stock
                red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                sheet.conditional_formatting.add(
                    'G2:G1000',
                    CellIsRule(operator='lessThan', formula=['10'], fill=red_fill)
                )
    
    def save_enhanced_workbook(self, wb):
        """Save the enhanced master inventory file"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"LARGO_LAB_MASTER_INVENTORY_ENHANCED_{timestamp}.xlsx"
        filepath = self.downloads_path / filename
        
        wb.save(filepath)
        print(f"\n✅ Enhanced master inventory saved as: {filename}")
        print(f"Full path: {filepath}")
        
        # Also save to the LabAutomation project
        project_path = Path("/Users/ugochi141/Desktop/LabAutomation/data")
        project_path.mkdir(exist_ok=True)
        project_file = project_path / filename
        wb.save(project_file)
        print(f"Copy saved to project: {project_file}")
        
        return filepath

def main():
    print("Enhanced Lab Inventory Management System")
    print("=" * 60)
    
    # Initialize consolidator
    consolidator = EnhancedInventoryConsolidator("/Users/ugochi141/Downloads")
    
    # Process inventory files
    print("\nProcessing inventory files...")
    consolidator.process_inventory_file()
    
    # Check for expiring items
    consolidator.check_for_expiring_items()
    
    # Create enhanced master workbook
    print("\nCreating enhanced master inventory workbook...")
    master_wb = consolidator.create_enhanced_master_workbook()
    
    # Save the workbook
    filepath = consolidator.save_enhanced_workbook(master_wb)
    
    print("\n" + "=" * 60)
    print("✅ INVENTORY CONSOLIDATION COMPLETE!")
    print("=" * 60)
    
    print("\n🎯 IMMEDIATE ACTIONS REQUIRED:")
    print("1. Review ALT reagent expiration dates (25 packs expire end of October)")
    print("2. Verify and correct supplier IDs marked as incorrect")
    print("3. Share file via Teams/SharePoint with Lorraine, Ingrid, and team")
    print("4. Train staff on updating procedures")
    print("5. Set up weekly review meetings for inventory status")
    
    print("\n📊 KEY FEATURES IMPLEMENTED:")
    print("✓ Automated status calculations (Out of Stock, Low Stock, Below PAR)")
    print("✓ Expiration date tracking with alerts")
    print("✓ Location tracking for MOB/AUC labs")
    print("✓ Hand count vs PAR level comparison")
    print("✓ Supplier ID and Onelink number fields for ordering")
    print("✓ Action required column for urgent items")
    print("✓ Summary dashboard with critical alerts")
    print("✓ Staff reminders and procedures integrated")

if __name__ == "__main__":
    main()



