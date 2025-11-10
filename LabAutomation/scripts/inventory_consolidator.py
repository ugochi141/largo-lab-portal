#!/usr/bin/env python3
"""
Lab Inventory Management System Consolidator
Creates a comprehensive inventory management solution from existing Excel files
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os
from pathlib import Path

class InventoryConsolidator:
    def __init__(self, downloads_path):
        self.downloads_path = Path(downloads_path)
        self.inventory_files = []
        self.master_inventory = {}
        
    def find_inventory_files(self):
        """Find all relevant inventory Excel files"""
        patterns = ['*inventory*.xlsx', '*Inventory*.xlsx']
        for pattern in patterns:
            self.inventory_files.extend(list(self.downloads_path.glob(pattern)))
        
        # Exclude temporary files
        self.inventory_files = [f for f in self.inventory_files if not f.name.startswith('~$')]
        print(f"Found {len(self.inventory_files)} inventory files")
        
    def read_shanthi_nate_file(self):
        """Read the main inventory file with hand counts"""
        main_file = self.downloads_path / "Inventory List Largo 03MAR25_Shanthi and Nate B_09092025.xlsx"
        if not main_file.exists():
            print(f"Main file not found: {main_file}")
            return
            
        wb = openpyxl.load_workbook(main_file, data_only=True)
        
        for sheet_name in wb.sheetnames:
            print(f"Processing sheet: {sheet_name}")
            sheet = wb[sheet_name]
            
            # Skip empty sheets
            if sheet.max_row < 2:
                continue
                
            # Extract header row to identify columns
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value if cell.value else "")
            
            # Process data rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue
                    
                item_data = {}
                for idx, value in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        item_data[headers[idx]] = value
                
                # Create unique key for the item
                if 'DESCRIPTION' in item_data and item_data['DESCRIPTION']:
                    key = f"{sheet_name}_{item_data.get('DESCRIPTION', '')}"
                    self.master_inventory[key] = {
                        'category': sheet_name,
                        'data': item_data
                    }
        
        print(f"Loaded {len(self.master_inventory)} items from main file")
        
    def create_master_workbook(self):
        """Create the comprehensive master inventory workbook"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define categories
        categories = ['CHEMISTRY', 'HEMATOLOGY', 'URINALYSIS', 'KITS', 'MISCELLANEOUS', 'SUMMARY']
        
        # Define standard columns for inventory tracking
        standard_columns = [
            'DESCRIPTION',
            'MFR#/CAT#',
            'MATERIAL# (KAISER#/OLID)',
            'SUPPLIER ID',
            'ONELINK NUMBER',
            'HAND COUNT',
            'REQ QTY',
            'EXPIRATION DATE',
            'LOT NUMBER',
            'LOCATION (MOB/AUC)',
            'LAST UPDATED',
            'UPDATED BY',
            'NOTES'
        ]
        
        # Create sheets for each category
        for category in categories:
            sheet = wb.create_sheet(title=category)
            
            # Add headers with formatting
            for col, header in enumerate(standard_columns, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Adjust column widths
            column_widths = {
                'A': 40,  # DESCRIPTION
                'B': 20,  # MFR#/CAT#
                'C': 25,  # MATERIAL#
                'D': 15,  # SUPPLIER ID
                'E': 20,  # ONELINK NUMBER
                'F': 12,  # HAND COUNT
                'G': 10,  # REQ QTY
                'H': 15,  # EXPIRATION DATE
                'I': 15,  # LOT NUMBER
                'J': 15,  # LOCATION
                'K': 15,  # LAST UPDATED
                'L': 20,  # UPDATED BY
                'M': 30   # NOTES
            }
            
            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width
        
        # Populate data from master inventory
        for item_key, item_info in self.master_inventory.items():
            category = item_info['category'].upper()
            if category not in categories:
                category = 'MISCELLANEOUS'
                
            sheet = wb[category]
            row = sheet.max_row + 1
            
            data = item_info['data']
            
            # Map data to standard columns
            sheet.cell(row=row, column=1, value=data.get('DESCRIPTION', ''))
            sheet.cell(row=row, column=2, value=data.get('MFR#/CAT#', ''))
            sheet.cell(row=row, column=3, value=data.get('MATERIAL# (KAISER#/OLID)', ''))
            sheet.cell(row=row, column=4, value=data.get('SUPPLIER ID', ''))
            sheet.cell(row=row, column=5, value=data.get('ONELINK NUMBER', ''))
            sheet.cell(row=row, column=6, value=data.get('Hand Count', ''))
            sheet.cell(row=row, column=7, value=data.get('Req QTY', ''))
            sheet.cell(row=row, column=11, value=datetime.now().strftime('%Y-%m-%d'))
            sheet.cell(row=row, column=12, value='Initial Import')
            
            # Highlight items with low counts
            if isinstance(data.get('Hand Count', 0), (int, float)) and data.get('Hand Count', 0) < 10:
                for col in range(1, 14):
                    sheet.cell(row=row, column=col).fill = PatternFill(
                        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                    )
        
        # Create Summary sheet
        self.create_summary_sheet(wb)
        
        # Add data validation and formulas
        self.add_validation_and_formulas(wb)
        
        return wb
    
    def create_summary_sheet(self, wb):
        """Create a summary dashboard sheet"""
        summary = wb['SUMMARY']
        
        # Title
        summary.merge_cells('A1:F1')
        title_cell = summary['A1']
        title_cell.value = "LARGO LAB INVENTORY SUMMARY DASHBOARD"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            ['Category', 'Total Items', 'Items < 10', 'Items Expiring Soon', 'Last Updated'],
            ['CHEMISTRY', '=COUNTA(CHEMISTRY!A:A)-1', '=COUNTIF(CHEMISTRY!F:F,"<10")', '', '=TODAY()'],
            ['HEMATOLOGY', '=COUNTA(HEMATOLOGY!A:A)-1', '=COUNTIF(HEMATOLOGY!F:F,"<10")', '', '=TODAY()'],
            ['URINALYSIS', '=COUNTA(URINALYSIS!A:A)-1', '=COUNTIF(URINALYSIS!F:F,"<10")', '', '=TODAY()'],
            ['KITS', '=COUNTA(KITS!A:A)-1', '=COUNTIF(KITS!F:F,"<10")', '', '=TODAY()'],
            ['MISCELLANEOUS', '=COUNTA(MISCELLANEOUS!A:A)-1', '=COUNTIF(MISCELLANEOUS!F:F,"<10")', '', '=TODAY()']
        ]
        
        for row_idx, row_data in enumerate(headers, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Instructions section
        summary.merge_cells('A10:F10')
        instructions_title = summary['A10']
        instructions_title.value = "INSTRUCTIONS FOR USE"
        instructions_title.font = Font(bold=True, size=14)
        
        instructions = [
            "1. Update HAND COUNT column when performing inventory checks",
            "2. Add EXPIRATION DATE for reagents and time-sensitive items",
            "3. Use LOCATION column to specify MOB or AUC lab",
            "4. Add notes about discrepancies or special handling in NOTES column",
            "5. Items with count < 10 are highlighted in yellow",
            "6. Review and update SUPPLIER ID and ONELINK NUMBER for accuracy",
            "7. Save file after each update and share via Teams/SharePoint"
        ]
        
        for idx, instruction in enumerate(instructions, 12):
            summary.cell(row=idx, column=1, value=instruction)
    
    def add_validation_and_formulas(self, wb):
        """Add data validation and conditional formatting"""
        from openpyxl.worksheet.datavalidation import DataValidation
        
        # Location validation
        location_dv = DataValidation(
            type="list",
            formula1='"MOB,AUC,BOTH"',
            allow_blank=True
        )
        location_dv.error = 'Please select MOB, AUC, or BOTH'
        location_dv.errorTitle = 'Invalid Location'
        
        for sheet_name in ['CHEMISTRY', 'HEMATOLOGY', 'URINALYSIS', 'KITS', 'MISCELLANEOUS']:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                location_dv.add(f'J2:J1000')
                sheet.add_data_validation(location_dv)
    
    def save_master_inventory(self, wb):
        """Save the master inventory file"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"LARGO_LAB_MASTER_INVENTORY_{timestamp}.xlsx"
        filepath = self.downloads_path / filename
        
        wb.save(filepath)
        print(f"\nMaster inventory saved as: {filename}")
        print(f"Full path: {filepath}")
        
        # Also save to the LabAutomation project
        project_path = Path("/Users/ugochi141/Desktop/LabAutomation/data")
        project_path.mkdir(exist_ok=True)
        project_file = project_path / filename
        wb.save(project_file)
        print(f"Copy saved to project: {project_file}")
        
        return filepath

def main():
    print("Lab Inventory Consolidation System")
    print("=" * 50)
    
    # Initialize consolidator
    consolidator = InventoryConsolidator("/Users/ugochi141/Downloads")
    
    # Find inventory files
    consolidator.find_inventory_files()
    
    # Read the main file with hand counts
    consolidator.read_shanthi_nate_file()
    
    # Create master workbook
    print("\nCreating master inventory workbook...")
    master_wb = consolidator.create_master_workbook()
    
    # Save the workbook
    filepath = consolidator.save_master_inventory(master_wb)
    
    print("\n✅ Inventory consolidation complete!")
    print("\nNext steps:")
    print("1. Review the generated file for accuracy")
    print("2. Update any missing supplier IDs or material numbers")
    print("3. Add expiration dates for time-sensitive items")
    print("4. Share via Teams/SharePoint for team access")
    print("5. Train staff on updating procedures")

if __name__ == "__main__":
    main()



