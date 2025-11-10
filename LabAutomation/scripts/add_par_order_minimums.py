#!/usr/bin/env python3
"""
Add PAR levels and order minimums to all inventory items
Based on usage patterns and critical nature of items
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import shutil
from datetime import datetime

def add_par_and_minimums():
    """Add realistic PAR levels and order minimums to inventory"""
    
    # Load the existing detailed inventory
    filepath = Path("/Users/ugochi141/Downloads/LARGO_LAB_DETAILED_INVENTORY_SYSTEM.xlsx")
    if not filepath.exists():
        filepath = Path("/Users/ugochi141/Desktop/LabAutomation/data/inventory/LARGO_LAB_DETAILED_INVENTORY_SYSTEM.xlsx")
    
    # Create backup
    backup_path = filepath.parent / f"backup_{filepath.name}"
    shutil.copy2(filepath, backup_path)
    
    # Load workbook
    wb = openpyxl.load_workbook(filepath)
    
    # Define PAR levels and order minimums by category and item type
    par_data = {
        'CHEMISTRY': {
            # High volume tests
            'Glucose': {'par': 15, 'min': 8, 'order_min': 10, 'max': 30},
            'Creatinine': {'par': 12, 'min': 6, 'order_min': 8, 'max': 25},
            'ALT': {'par': 8, 'min': 4, 'order_min': 6, 'max': 15, 'note': 'CHECK EXPIRATION!'},
            'ASTP': {'par': 8, 'min': 4, 'order_min': 6, 'max': 15},
            'Total Bilirubin': {'par': 6, 'min': 3, 'order_min': 4, 'max': 12},
            'Albumin': {'par': 8, 'min': 4, 'order_min': 6, 'max': 15},
            'Calcium': {'par': 8, 'min': 4, 'order_min': 6, 'max': 15},
            'Alkaline Phosphatase': {'par': 6, 'min': 3, 'order_min': 4, 'max': 12},
            'Total Protein': {'par': 6, 'min': 3, 'order_min': 4, 'max': 12},
            
            # Lower volume tests
            'Direct Bilirubin': {'par': 4, 'min': 2, 'order_min': 3, 'max': 8},
            'Digoxin': {'par': 2, 'min': 1, 'order_min': 2, 'max': 4},
            'Acetaminophen': {'par': 3, 'min': 1, 'order_min': 2, 'max': 6},
            'Salicylate': {'par': 3, 'min': 1, 'order_min': 2, 'max': 6},
            
            # ISE Electrodes (replace regularly)
            'ISE': {'par': 3, 'min': 2, 'order_min': 2, 'max': 5},
            'Electrode': {'par': 3, 'min': 2, 'order_min': 2, 'max': 5},
            
            # Maintenance solutions
            'Wash': {'par': 6, 'min': 3, 'order_min': 4, 'max': 12},
            'Diluent': {'par': 8, 'min': 4, 'order_min': 6, 'max': 15},
            'Cleaning': {'par': 4, 'min': 2, 'order_min': 3, 'max': 8}
        },
        'HEMATOLOGY': {
            # Controls (daily use)
            'Control': {'par': 3, 'min': 2, 'order_min': 2, 'max': 6},
            
            # Reagents (high volume)
            'Cellpack': {'par': 10, 'min': 5, 'order_min': 6, 'max': 20},
            'Stromatolyser': {'par': 6, 'min': 3, 'order_min': 4, 'max': 12},
            'Sulfolyser': {'par': 4, 'min': 2, 'order_min': 3, 'max': 8},
            'Cellclean': {'par': 4, 'min': 2, 'order_min': 3, 'max': 8},
            
            # Staining
            'Stain': {'par': 3, 'min': 2, 'order_min': 2, 'max': 6},
            'Immersion Oil': {'par': 2, 'min': 1, 'order_min': 1, 'max': 4}
        },
        'URINALYSIS': {
            # Test strips (high volume)
            'Test Strips': {'par': 15, 'min': 8, 'order_min': 10, 'max': 30},
            
            # Controls
            'Control': {'par': 3, 'min': 2, 'order_min': 2, 'max': 6},
            
            # Collection supplies
            'Collection Cups': {'par': 6, 'min': 3, 'order_min': 4, 'max': 12},
            'Transfer Tubes': {'par': 15, 'min': 8, 'order_min': 10, 'max': 30},
            
            # Microscopy
            'Slides': {'par': 10, 'min': 5, 'order_min': 5, 'max': 20},
            'Cover Slips': {'par': 10, 'min': 5, 'order_min': 5, 'max': 20}
        },
        'KITS': {
            # MEDTOX (special handling)
            'MEDTOX': {'par': 5, 'min': 3, 'order_min': 3, 'max': 10, 'note': 'LOG IN CERNER!'},
            
            # Coagulation controls
            'PT/INR': {'par': 3, 'min': 2, 'order_min': 2, 'max': 6},
            'PTT': {'par': 3, 'min': 2, 'order_min': 2, 'max': 6},
            
            # Rapid tests
            'COVID': {'par': 8, 'min': 4, 'order_min': 5, 'max': 15},
            'Flu': {'par': 5, 'min': 3, 'order_min': 3, 'max': 10},
            'Strep': {'par': 5, 'min': 3, 'order_min': 3, 'max': 10},
            'Pregnancy': {'par': 5, 'min': 3, 'order_min': 3, 'max': 10}
        },
        'MISCELLANEOUS': {
            # PPE (high usage)
            'Gloves': {'par': 30, 'min': 15, 'order_min': 20, 'max': 60},
            
            # Thermometers
            'Thermometer': {'par': 3, 'min': 2, 'order_min': 2, 'max': 5},
            
            # Pipettes
            'Pipette': {'par': 3, 'min': 2, 'order_min': 2, 'max': 5},
            
            # Safety
            'Sharps': {'par': 3, 'min': 2, 'order_min': 2, 'max': 6},
            'Biohazard': {'par': 5, 'min': 3, 'order_min': 3, 'max': 10},
            
            # Office supplies
            'Paper': {'par': 10, 'min': 5, 'order_min': 5, 'max': 20},
            'Log Book': {'par': 3, 'min': 2, 'order_min': 2, 'max': 5}
        }
    }
    
    # Process each category sheet
    for sheet_name in ['CHEMISTRY', 'HEMATOLOGY', 'URINALYSIS', 'KITS', 'MISCELLANEOUS']:
        if sheet_name not in wb.sheetnames:
            continue
            
        sheet = wb[sheet_name]
        category_pars = par_data.get(sheet_name, {})
        
        # Find column indices (1-based)
        par_col = 10  # PAR LEVEL
        min_col = 11  # MIN STOCK
        max_col = 12  # MAX STOCK
        reorder_col = 15  # REORDER POINT
        notes_col = 29  # NOTES
        
        # Update header row to emphasize these columns
        sheet.cell(row=1, column=par_col).fill = PatternFill(
            start_color="FFD700", end_color="FFD700", fill_type="solid"
        )
        sheet.cell(row=1, column=min_col).fill = PatternFill(
            start_color="FFD700", end_color="FFD700", fill_type="solid"
        )
        sheet.cell(row=1, column=reorder_col).fill = PatternFill(
            start_color="FFD700", end_color="FFD700", fill_type="solid"
        )
        
        # Process each item row
        for row in range(2, sheet.max_row + 1):
            desc_cell = sheet.cell(row=row, column=2)  # DESCRIPTION column
            if not desc_cell.value:
                continue
                
            description = str(desc_cell.value).upper()
            
            # Find matching PAR data
            par_info = None
            for key, data in category_pars.items():
                if key.upper() in description:
                    par_info = data
                    break
            
            # Apply default if no match
            if not par_info:
                # Default PAR levels by category
                defaults = {
                    'CHEMISTRY': {'par': 5, 'min': 3, 'order_min': 3, 'max': 10},
                    'HEMATOLOGY': {'par': 4, 'min': 2, 'order_min': 3, 'max': 8},
                    'URINALYSIS': {'par': 6, 'min': 3, 'order_min': 4, 'max': 12},
                    'KITS': {'par': 4, 'min': 2, 'order_min': 3, 'max': 8},
                    'MISCELLANEOUS': {'par': 5, 'min': 3, 'order_min': 3, 'max': 10}
                }
                par_info = defaults.get(sheet_name, {'par': 5, 'min': 3, 'order_min': 3, 'max': 10})
            
            # Update cells with PAR data
            sheet.cell(row=row, column=par_col, value=par_info['par'])
            sheet.cell(row=row, column=min_col, value=par_info['min'])
            sheet.cell(row=row, column=max_col, value=par_info['max'])
            sheet.cell(row=row, column=reorder_col, value=par_info['order_min'])
            
            # Highlight cells with low PAR
            if par_info['par'] <= 3:
                sheet.cell(row=row, column=par_col).fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                )
            
            # Add notes if specified
            if 'note' in par_info:
                current_note = sheet.cell(row=row, column=notes_col).value or ''
                if par_info['note'] not in str(current_note):
                    new_note = f"{current_note} {par_info['note']}" if current_note else par_info['note']
                    sheet.cell(row=row, column=notes_col, value=new_note)
                    sheet.cell(row=row, column=notes_col).font = Font(bold=True, color="FF0000")
            
            # Special handling for ALT reagents
            if 'ALT' in description and 'REAGENT' in description:
                sheet.cell(row=row, column=par_col, value=8)
                sheet.cell(row=row, column=min_col, value=4)
                sheet.cell(row=row, column=reorder_col, value=6)
                sheet.cell(row=row, column=notes_col).value = "⚠️ 25 PACKS EXPIRE OCT 31! Order minimum 6 packs when below 4"
                
                # Highlight entire PAR section
                for col in [par_col, min_col, max_col, reorder_col]:
                    sheet.cell(row=row, column=col).fill = PatternFill(
                        start_color="FF6666", end_color="FF6666", fill_type="solid"
                    )
                    sheet.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
    
    # Update DASHBOARD with PAR summary
    if 'DASHBOARD' in wb.sheetnames:
        dashboard = wb['DASHBOARD']
        
        # Find a good spot to add PAR summary (around row 35)
        row = 35
        dashboard.merge_cells(f'A{row}:J{row}')
        dashboard[f'A{row}'] = "📊 PAR LEVEL & ORDER MINIMUM SUMMARY"
        dashboard[f'A{row}'].font = Font(bold=True, size=14)
        dashboard[f'A{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        row += 2
        par_summary = [
            "CRITICAL PAR LEVELS TO MONITOR:",
            "• HIGH VOLUME TESTS: PAR 12-15, Order when ≤8, Min order 10 packs",
            "• MEDIUM VOLUME TESTS: PAR 6-8, Order when ≤4, Min order 6 packs",
            "• LOW VOLUME TESTS: PAR 2-4, Order when ≤2, Min order 3 packs",
            "• ISE ELECTRODES: PAR 3, Order when ≤2, Min order 2 units (4-6 week life)",
            "• CONTROLS: PAR 3, Order when ≤2, Min order 2 boxes",
            "• PPE/GLOVES: PAR 20-30 boxes, Order when ≤15, Min order 20 boxes",
            "",
            "⚠️ SPECIAL ALERT - ALT REAGENTS:",
            "   Current: 25 packs (OVERSTOCKED)",
            "   PAR Level: 8 packs",
            "   Order Point: 4 packs",
            "   Min Order: 6 packs",
            "   ACTION: REDISTRIBUTE 17 EXCESS PACKS BEFORE OCT 31!"
        ]
        
        for line in par_summary:
            dashboard[f'A{row}'] = line
            dashboard.merge_cells(f'A{row}:J{row}')
            if line.startswith('⚠️'):
                dashboard[f'A{row}'].font = Font(bold=True, color="FF0000", size=12)
            elif line.startswith('   ACTION:'):
                dashboard[f'A{row}'].font = Font(bold=True, color="FF0000", size=14)
                dashboard[f'A{row}'].fill = PatternFill(
                    start_color="FFFF99", end_color="FFFF99", fill_type="solid"
                )
            row += 1
    
    # Save updated file
    new_filename = "LARGO_LAB_INVENTORY_WITH_PAR_LEVELS.xlsx"
    new_filepath = filepath.parent / new_filename
    wb.save(new_filepath)
    
    # Also save to project
    project_path = Path("/Users/ugochi141/Desktop/LabAutomation/data/inventory") / new_filename
    wb.save(project_path)
    
    print(f"✅ Updated inventory with PAR levels saved as: {new_filename}")
    print(f"   Downloads: {new_filepath}")
    print(f"   Project: {project_path}")
    
    return new_filepath

def main():
    print("Adding PAR Levels and Order Minimums to Inventory")
    print("=" * 60)
    
    filepath = add_par_and_minimums()
    
    print("\n" + "=" * 60)
    print("📊 PAR LEVELS ADDED:")
    print("=" * 60)
    print("✓ PAR levels based on test volume and criticality")
    print("✓ Minimum stock levels to prevent outages")
    print("✓ Order minimums for efficient purchasing")
    print("✓ Maximum stock to prevent overstocking")
    print("✓ Special alerts for ALT reagent overstocking")
    print("✓ Color coding for low PAR items")
    print("✓ Dashboard summary of PAR guidelines")
    
    print("\n🎯 KEY PAR LEVELS:")
    print("• Glucose/Creatinine: PAR 12-15 (high volume)")
    print("• ALT/AST: PAR 8 (but currently 25 - OVERSTOCKED!)")
    print("• Controls: PAR 3 (daily use)")
    print("• PPE Gloves: PAR 20-30 boxes")
    print("• Rapid Tests: PAR 5-8 kits")
    
    print("\n⚠️ IMMEDIATE ACTION:")
    print("ALT Reagents: 25 packs on hand vs PAR of 8")
    print("MUST redistribute 17 excess packs before expiration!")

if __name__ == "__main__":
    main()



